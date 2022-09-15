__author__ = "Jeremy Leese | jeremy.leese@thomsonreuters.com | Technology - Product Eng - Content Platforms"
__copyright__ = "Copyright© (C) 2022 Thomson Reuters. All Rights Reserved."
__version__ = "0.19"

from operator import truediv
import ldap 
from ldif import LDIFParser,LDIFWriter
import ldap.modlist as modlist
import util
import json
import re

import boto3
from botocore.exceptions import ClientError
import openpyxl

import migmaps

ldap_py = None
LOG = None
CONN = None
resource_list=[]
SECRETS_MANAGER_SERVICE = "secretsmanager"
REGION_NAME = 'us-east-1'

##
# BEGIN DEF return client for specified AWS region and service
##
def get_client(regionName, serviceName):
    session = boto3.session.Session()
    client = session.client(service_name=serviceName, region_name=regionName)
    
    return client

# END DEF get_client

'''
##
# get_secrets_list()
##
'''
def get_secrets_list(client):

    secrets_list = None
    next_token = None

    try:
        response = client.list_secrets(MaxResults=100)
        secrets_list = response.get('SecretList')
        next_token = response.get('NextToken')

        while(next_token):
            response = client.list_secrets(MaxResults=100, NextToken=next_token)
            secrets_list.extend(response.get('SecretList'))
            next_token = response.get('NextToken')

        return secrets_list
    
    except ClientError as e:
        raise Exception("boto3 client error in get_secrets: " + e.__str__())
    except Exception as e:
        raise Exception("Unexpected error in get_secrets: " + e.__str__()) 

##
# modify_ldap(l)
##
def modify_ldap(dn, old, new):
   
   ''' DEBUG
   print()
   print("DN - " + dn)
   print("OLD - ", old)
   print("NEW - ", new)
   print()
   '''
   
   # Open a connection
   l = ldap.initialize('ldap://localhost:389')

   # Bind/authenticate with a user with apropriate rights to add objects
   l.simple_bind_s("cn=root,cn=qc,ou=novusaws,o=westgroup.com", "&comPP3tellymanGGdoF")

   # Convert place-holders for modify-operation using modlist-module
   ldif = modlist.modifyModlist(old,new)
   print(ldif)

   # Do the actual modification 
   l.modify_s(dn,ldif)

   # disconnect and free resources when done
   l.unbind_s() 
 
'''
### bucket_fix()
'''
def bucket_fix():
   dn = 'cn=metadoc.cms.01,cn=Resources,cn=qc,ou=novusaws,o=westgroup.com'

   # A dict to help build the "body" of the object
   attrs = {}
   attrs['objectclass'] = [b'top',b'wgDataStore']
   attrs['cn'] = [b'metadoc.cms.01.dev']

   # Convert our dict to nice syntax for the add-function using modlist-module
   ldif = modlist.addModlist(attrs)

   mod_attrs = [( ldap.MOD_ADD, 'cn', b'dev.bucket')]
   add_dn = 'cn=metadoc.cms.01.dev,' + dn 
   #add_dn = dn

   ldap_py.add_s(add_dn, ldif)
   ldap_py.modify_s(add_dn,mod_attrs)

   old = {'wgResourceName': [b'cmsr.04a.update']}
   new = {'wgResourceName': [b'metadoc.cms.04']}
   dn = 'cn=metadoc.cms.04,cn=Resources,cn=qc,ou=novusaws,o=westgroup.com'

   ldif = modlist.modifyModlist(old,new)
   print(ldif)

   # Do the actual modification 
   ldap_py.modify_s(dn,ldif)



'''
# update_hosting_metadata
'''
def update_hosting_metadata(dn, entry):
   

   wg_arg = entry.get('wgArgument')

   if wg_arg:

      if str(wg_arg).lower().find('env=qc') >= 0:
         
         print()
         print('dn:', dn)

         found = False 

         if str(wg_arg).find('env=qc') >= 0:
            new_wg_arg = str(wg_arg).replace('env=qc', 'env=novusaws:qc')
            found = True
         elif str(wg_arg).find('ENV=qc') >= 0:
            new_wg_arg = str(wg_arg).replace('ENV=qc', 'env=novusaws:qc')
            found = True
         elif str(wg_arg).find('env=QC') >= 0:
            new_wg_arg = str(wg_arg).replace('env=QC', 'env=novusaws:qc')
            found = True
         elif str(wg_arg).find('ENV=QC') >= 0:
            new_wg_arg = str(wg_arg).replace('ENV=QC', 'env=novusaws:qc')
            found = True

         if found:
            print()
            print(type(wg_arg[0]), 'wg_arg[0]:', wg_arg[0])
            print(type(wg_arg), 'wg_arg:', wg_arg)
            print(type(new_wg_arg), 'new_wg_arg:', new_wg_arg)
            new_wg_arg = new_wg_arg[3:-2]

            old_value = {'wgArgument': wg_arg}
            new_value = {'wgArgument': [str.encode(new_wg_arg)]}

            print()
            print('old_value', type(old_value), old_value)
            print('new_value', type(new_value), new_value)

            modlist = ldap.modlist.modifyModlist(old_value, new_value)
            print(modlist)
            ldap_py.modify_s(dn, modlist)

         else:
            print('dev env not found:', wg_arg)
         print()

'''
# BEGIN get_migration_data()
'''
def get_migration_data():

   #filename = "C:\\AWS\\py\\qc_wmd.xlsx"
   filename = "C:\\AWS\\py\\qc_mig_map.xlsx" 
   wb = openpyxl.load_workbook(filename)

   ws = wb.active

   max_col = ws.max_column
   max_row = ws.max_row

   migration_list = []
   migration_dict = {}

   for i in range(2, max_row):

      # add xcel rows that have been populated with AWS migration data     
      if ws.cell(row = i, column = 3).value is not None:
         #print('cell:', ws.cell(row = i, column = 3).value)
         migration_dict = {
               "legacy_resource": ws.cell(row = i, column = 1).value,
               "legacy_cn": ws.cell(row = i, column = 2).value,
               "aws_resource": ws.cell(row = i, column = 3).value,
               "aws_ds_ptr": ws.cell(row = i, column = 4).value,
               "legacy_ds_ptr": ws.cell(row = i, column = 5).value,
               "ldap_dn": ws.cell(row = i, column = 6).value
         }

         migration_list.append(migration_dict)

   return migration_list

'''
# migrate_accessible_datastores()
'''
def migrate_accessible_datastores(commit):

   baseDN = "cn=qc,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   migration_list = get_migration_data()

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0
      print('COMMIT=', commit)

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]

         #print('dn:', dn)

         for acc_ds in migration_list:

            ldap_dn = acc_ds.get("ldap_dn")
            #print('\tldap_dn:', ldap_dn)
            if ldap_dn.find(dn) >= 0:
               count = count +1
               #print()
               #print('COUNT:', count)
               migrate_data = acc_ds
            
               leg_res = migrate_data.get('legacy_resource')
               aws_res = migrate_data.get('aws_resource')
               leg_dsp = migrate_data.get('legacy_ds_ptr')
               aws_dsp = migrate_data.get('aws_ds_ptr')

               #print('res data->', leg_res, ':', aws_res)
               #print('dsp data->', leg_dsp, ':', aws_dsp)

               old_entry = {"wgResourceName": [leg_res.encode('utf-8')]}
               new_entry = {"wgResourceName": [aws_res.encode('utf-8')]}
               res_modlist = ldap.modlist.modifyModlist(old_entry, new_entry)

               old_entry = {"wgDataStorePtr": [leg_dsp.encode('utf-8')]}
               new_entry = {"wgDataStorePtr": [aws_dsp.encode('utf-8')]}
               dsp_modlist = ldap.modlist.modifyModlist(old_entry, new_entry)
               
               if commit:
                  
                  print(file=LOG)
                  print('Updating:', ldap_dn)
                  print('legacy data->', leg_res, ':', leg_dsp, file=LOG)
                  print('res_modlist->', res_modlist, file=LOG )
                  print('dsp_modlist->', dsp_modlist, file=LOG )

                  #print(res_modlist)
                  #print(dsp_modlist)

                  ldap_py.modify_s(ldap_dn, res_modlist)
                  ldap_py.modify_s(ldap_dn, dsp_modlist)

               else:
                  print()
                  print('legacy data->', leg_res, ':', leg_dsp)
                  print(res_modlist)
                  print(dsp_modlist)

               continue

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

'''
# remove_ro_db_servers()
'''
def remove_ro_db_servers(commit):

   baseDN = "cn=DatabaseServers,cn=qc,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgDatabaseServer)"

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count1 = 0
      count2 = 0
      not_started = 'not started'
      started = 'started'

      print('COMMIT=', commit)

      for entry in result_set:
         count1 = count1 + 1

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         
         if cn.find('-ro-') >= 0:
            count2 = count2 + 1
            if(commit):
               print('COMMIT DELETE:', dn)
               ldap_py.delete(dn)
            else:
               print('NO COMMIT DELETE:', dn)
         
         if cn.find('z-') >= 0 or cn.find('b-') >= 0:
            print()
            print('cn:', cn)
            #print(entry_dict)
            print('status:', entry_dict.get('wgStatus')[0])

            ldif = modlist.modifyModlist({'wgStatus': entry_dict['wgStatus'][0]},{'wgStatus': [not_started.encode()]})
            print(ldif) 
            ldap_py.modify_s(dn,ldif)

         else:
            print()
            print('cn:', cn)
            #print(entry_dict)
            print('status:', entry_dict.get('wgStatus')[0])

            ldif = modlist.modifyModlist({'wgStatus': entry_dict['wgStatus'][0]},{'wgStatus': [started.encode()]})
            print(ldif) 
            ldap_py.modify_s(dn,ldif)

      print(count1)
      print(count2)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

'''
# add_resource      
'''
def add_resource(resource, buckets, commit):

   dn = 'cn=' + resource + ',cn=Resources,cn=client,ou=novusaws,o=westgroup.com'

   # A dict to help build the "body" of the object
   attrs = {}
   attrs['objectclass'] = [b'top',b'wgResource']
   attrs['cn'] = [resource.encode()]
   attrs['wgCurrentBuckets'] = [buckets.encode('utf-8')]
   attrs['wgNewBuckets'] = [buckets.encode('utf-8')]

   # Convert our dict to nice syntax for the add-function using modlist-module
   ldif = modlist.addModlist(attrs)

   print()
   print('dn', dn)
   print('ldif: ', ldif)

   if(commit):

      try:
         ldap_py.add_s(dn, ldif)
      except Exception as e:
         print()
         print(e)

'''
# add_datastores()
'''
def update_datastores(log):
   print('update_datastores(): BEGIN')
   print('update_datastores(): BEGIN', file=log)

   baseDN = "cn=Resources,cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgResource)"

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count1 = 0

      for entry in result_set:
         count1 = count1 + 1

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')

         print()
         print('cn:', cn)
         
         '''
         if cn.find('read') >= 0:
            if db_host.find('-ro-') >= 0:
               print()
               print('dn:', dn)
               print('cn:', cn)
               print('wgDatabaseHost:', db_host)
               print(entry_dict)

               db_host = db_host.replace('-ro', '')

               ldif = modlist.modifyModlist({'wgDatabaseHost': entry_dict['wgDatabaseHost'][0]},{'wgDatabaseHost': [db_host.encode()]})
               print(ldif) 
               ldap_py.modify_s(dn,ldif)
         '''

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

'''
END get_ldap_resource_dict()
'''

def add_bucket_attributes(dn, entry, commit):

   # safety mechanism
   #commit = False

   bucketed_resources_4 = ['persist']
   bucketed_resources_8 = ['dlc.01','dlc.02']
   bucketed_resources_32 = ['norm.alpha.01','norm.bpi3.01','norm.cciunitcoomp.01','norm.connect.01','norm.connect.02','norm.generic.01','norm.loadreg.01','norm.bpi2.01','norm.perf1.01','nort.perf1.01','nims.ccscalertr.01','nims.ccscalertr.02','nims.connect.01','nims.connect.02','metadoc.cms.01','metadoc.cms.02','metadoc.cms.03','metadoc.cms.04']
   buckets_32 = "32"
   buckets_16 = "16"
   buckets_8 = "8"
   buckets_4 = "4"

   if str(entry['objectClass'][1]).find('wgDataStore') >= 0:

      cn = str(entry['cn'][0], 'utf-8')

      #print('DN:', dn)
      print('CN', cn)
      #print('ENTRY:', entry)

      if cn in bucketed_resources_4:
         buckets = buckets_4
      elif cn in bucketed_resources_8:
         buckets = buckets_8
      elif cn in bucketed_resources_32:
         buckets = buckets_32
      else:
         return # bad programmer is bad

      mod_attrs = [( ldap.MOD_ADD, 'wgCurrentBuckets', buckets.encode())]
      print('mod_attrs1:',mod_attrs)

      if(commit):
         ldap_py.modify_s(dn,mod_attrs)

      mod_attrs = [( ldap.MOD_ADD, 'wgNewBuckets', buckets.encode())]
      print('mod_attrs2:',mod_attrs)

      if(commit):
         ldap_py.modify_s(dn,mod_attrs)



'''
# audit_datastore_ptrs()
'''
def audit_datastore_ptrs(ldap_py, log):
   print('audit_datastore_ptrs(): BEGIN')
   #print('audit_datastore_ptrs(): BEGIN', file=log)


   #out_log = open('qc_uds_update.log3', 'w')
   #baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   baseDN = "cn=prod,ou=novusaws,o=westgroup.com"
   #baseDN = "cn=qc,ou=novusaws,o=westgroup.com"
   #baseDN = "cn=dev,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   """ retrieve all attributes - see documentation for more options """
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   count = 0

   #resource_dict = util.get_resource_dict()

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []
      
      print()
   
      print('dsp', '|\t', 'cn', '|\t', 'rsc', '|\t', 'dn' )
      #print('dsp', ':\t', 'cn', ':\t', 'rsc', ':\t', 'dn', file=log)

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0

      for entry in result_set:
         

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')
      
         #print(dn)

         if 'resources' in dsp:

            count = count + 1

            if 'master' in dsp.lower():      
               #if rsc in rsc_dict:
               print(dsp, '| ', cn, '|', rsc, '|', dn )
               print(dsp, cn, rsc, dn, file=log)
            elif 'slave' in dsp.lower():      
               #if rsc in rsc_dict:
               print(dsp, '| ', cn, '|', rsc, '|', dn )
               print(dsp, cn, rsc, dn, file=log)
            else:
               print(dsp, '| ', cn, '|', rsc, '|', dn )
               print(dsp, cn, rsc, dn, file=log)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print(count, 'DSP entries found.')
      print('audit_datastore_ptrs(): END')
      #print('audit_datastore_ptrs(): END', file=log)


class ParserStub(LDIFParser):
    def __init__(self, *args, **kwargs):
        self.records = []
        super().__init__(*args, **kwargs)

    def handle(self, dn, entry):
        self.records.append((dn, entry))

'''
# 
'''
def add_bucket_datastores_std(main_resource, start_bucket, buckets, db_override, commit):

   #failsafe
   #commit = False

   ds_attrs = {}
   ds_attrs['objectclass'] = [b'top',b'wgDataStore']
   ds_attrs['wgAWSRegion'] = [b'us-east-1']
   ds_attrs['wgDataStoreType'] = [b'POSTGRES NO_VARCHARS STREAMS STATEMENTCACHESIZE=20']
   ds_attrs['wgJDBCDriverName'] = [b'org.postgresql.Driver']

   update_ar = True
   update_au = True
   update_zr = True
      
   for i in range(buckets):

      bucket_num = '.' + str(i + start_bucket)
      bucket_resource = main_resource + bucket_num

      if(update_ar):
         '''
            primary read datastore
         '''
         cn = main_resource + 'a.read'

         # allow override for non-conventional names. usually buckets split across resource names.
         if db_override is None:
            db_name = main_resource.replace('.', '-')
         else:
            db_name = db_override
         
         pri_db_name = db_name + 'a'

         dn = 'cn=' + cn + ',cn='
         dn = dn + main_resource + bucket_num + ',cn='
         dn = dn + main_resource + ',cn=Resources,cn=client,ou=novusaws,o=westgroup.com'

         cluster_name = db_name.replace('-', '')
         cluster_name = 'pg' + cluster_name + 'aq'
         
         secret_name = 'a207947-novus-' + pri_db_name + '-db-read-qa-us-east-1'
         db_host = 'a207947-' + pri_db_name + '-qa-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com'
         jdbc_url = 'jdbc:postgresql://a207947-' + pri_db_name + '-qa-pg-cluster-us-east-1.cluster-ro-ccd4069y2ydl.us-east-1.rds.amazonaws.com:5432//' + cluster_name + '?sslmode=verify-ca&sslrootcert=/novus/novusqc/releases/db/rds-ca-2019-root.pem'
         
         ds_attrs['cn'] = [cn.encode()]
         ds_attrs['wgAWSSecretName'] = [secret_name.encode()]
         ds_attrs['wgDatabaseHost'] = [db_host.encode()]
         ds_attrs['wgJDBCConnectionURL'] = [jdbc_url.encode()]
         ds_attrs['wgReadOnly'] = [b'TRUE']
         ds_attrs['wgResourceName'] = [bucket_resource.encode()]
         
         ldif = modlist.addModlist(ds_attrs)
         #ldif = modlist.modifyModlist(ds_attrs, ds_attrs)
         print()
         print("Adding: ", dn)
         print(ldif)
         
         if commit:
            ldap_py.add_s(dn, ldif)
            #ldap_py.modify_s(dn, ldif)

      if(update_au):
         '''
            primary update datastore
         '''

         cn = main_resource + '.a.update'

         # allow override for non-conventional names. usually buckets split across resource names.
         if db_override is None:
            db_name = main_resource.replace('.', '-')
         else:
            db_name = db_override

         pri_db_name = db_name + 'a'

         dn = 'cn=' + cn + ',cn='
         dn = dn + main_resource + bucket_num + ',cn='
         dn = dn + main_resource + ',cn=Resources,cn=client,ou=novusaws,o=westgroup.com'

         cluster_name = db_name.replace('-', '')
         cluster_name = 'pg' + cluster_name + 'aq'
         
         secret_name = 'a207947-novus-' + pri_db_name + '-db-update-qa-us-east-1'
         db_host = 'a207947-' + pri_db_name + '-qa-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com'
         jdbc_url = 'jdbc:postgresql://a207947-' + pri_db_name + '-qa-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com:5432//' + cluster_name + '?sslmode=verify-ca&sslrootcert=/novus/novusqc/releases/db/rds-ca-2019-root.pem'
         
         ds_attrs['cn'] = [cn.encode()]
         ds_attrs['wgAWSSecretName'] = [secret_name.encode()]
         ds_attrs['wgDatabaseHost'] = [db_host.encode()]
         ds_attrs['wgJDBCConnectionURL'] = [jdbc_url.encode()]
         ds_attrs['wgReadOnly'] = [b'FALSE']
         ds_attrs['wgResourceName'] = [bucket_resource.encode()]
         
         ldif = modlist.addModlist(ds_attrs)
         print()
         print("Adding: ", dn)
         print(ldif)
         
         if commit:
            ldap_py.add_s(dn, ldif)

      if(update_zr):
         '''
            secondary read datastore
         '''
         cn = main_resource + '.z.read'

         # allow override for resource names that don't match their db name.
         if db_override is None:
            db_name = main_resource.replace('.', '-')
         else:
            db_name = db_override

         sec_db_name = db_name + 'z'

         dn = 'cn=' + cn + ',cn='
         dn = dn + main_resource + bucket_num + ',cn='
         dn = dn + main_resource + ',cn=Resources,cn=client,ou=novusaws,o=westgroup.com'

         cluster_name = db_name.replace('-', '')
         cluster_name = 'pg' + cluster_name + 'zq'
         
         secret_name = 'a207947-novus-' + sec_db_name + '-db-read-qa-us-east-1'
         db_host = 'a207947-' + sec_db_name + '-qa-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com'
         jdbc_url = 'jdbc:postgresql://a207947-' + sec_db_name + '-qa-pg-cluster-us-east-1.cluster-ro-ccd4069y2ydl.us-east-1.rds.amazonaws.com:5432//' + cluster_name + '?sslmode=verify-ca&sslrootcert=/novus/novusqc/releases/db/rds-ca-2019-root.pem'
         
         ds_attrs['cn'] = [cn.encode()]
         ds_attrs['wgAWSSecretName'] = [secret_name.encode()]
         ds_attrs['wgDatabaseHost'] = [db_host.encode()]
         ds_attrs['wgJDBCConnectionURL'] = [jdbc_url.encode()]
         ds_attrs['wgReadOnly'] = [b'TRUE']
         ds_attrs['wgResourceName'] = [bucket_resource.encode()]
      
         ldif = modlist.addModlist(ds_attrs)
         print()
         print("Adding: ", dn)
         print(ldif)
         
         if commit:
            ldap_py.add_s(dn, ldif)

###
''' dsp_quick_scan begin '''      
###
def dsp_quick_scan(ldap_py):
   #baseDN = "cn=dev,ou=novusaws,o=westgroup.com"
   #baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
   result_set = []

   while 1:
      
      result_type, result_data = ldap_py.result(ldap_result_id, 0)

      if (result_data == []):
         break
      else:
         if result_type == ldap.RES_SEARCH_ENTRY:
            result_set.append(result_data)

   rscList = []
   count = 0

   for entry in result_set:

      dn = entry[0][0]
      entry_dict = entry[0][1]
      cn = entry_dict.get('cn')[0].decode('utf-8')
      dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
      rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')

      dsp_split = dsp.split('cn=') 
      base_rsc =  dsp_split[len(dsp_split) - 2]
      base_rsc = base_rsc[:len(base_rsc) - 1]

      if dsp.find('cn=resources') >= 0:
         count = count + 1
         print()
         print('dn:', dn)
         print('\tdsp:', dsp)
         print('\tbase_rsc:', base_rsc)

         if base_rsc not in rscList:
            rscList.append(base_rsc)

            for i in rscList:
               print(i)
   
   print()
   print(count)

   for i in rscList:
      print(i)

'''
# update_accessible_datastores()
'''
def update_accessible_datastores(commit):

   commit = False

   baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgAccessibleDatastore)"
   dsp_tail = ',cn=client,ou=novusaws,o=westgroup.com'

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            print('BREAK')
            break
         else:
            #print(result_data)
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0
      print('COMMIT=', commit)

      op_rsc = 'wlnv.rac5.n1'
      aws_rsc = 'doc.westlaw.03'

      for entry in result_set:
         dsp_suf = ''
         dn = entry[0][0]

         entry_dict = entry[0][1]
         dsp =  entry_dict['wgDataStorePtr'][0].decode('utf-8')
         cn = entry_dict['cn'][0].decode('utf-8')
         rsc = entry_dict['wgResourceName'][0].decode('utf-8')

         #print('ads:', entry_dict)
         #print('dsp:', dsp)
         #print('cn :', cn)
         

         if(dsp.find('resource') >=0 ):
            print(dn, dsp, rsc, file=log)
            #print(dn, dsp, rsc)

         if(rsc.find(op_rsc) >=0):
            
            print()
            print("FOUND")
            print()

            if(cn.find('master.u')):
               dsp_suf = 'a.update'
            elif(cn.find('master.r')):
               dsp_suf = 'a.read'
            elif(cn.find('slave.r')):
               dsp_suf = 'z.read'
            else:
               print()
               print('SUF NF:', cn)
               print()
               continue
            
            pos = dsp.find(',')
            new_dsp = dsp[pos:]
            new_rsc = aws_rsc + dsp_suf
            new_dsp = new_rsc + new_dsp
            
            print('new_rsc:', new_rsc)
            print('new_dsp:', new_dsp)

            old_entry1 = {"wgResourceName": [rsc.encode('utf-8')]}
            new_entry1 = {"wgResourceName": [new_rsc.encode('utf-8')]}
            rsc_modlist = ldap.modlist.modifyModlist(old_entry1, new_entry1)

            old_entry = {"wgDataStorePtr": [dsp.encode('utf-8')]}
            new_entry = {"wgDataStorePtr": [new_dsp.encode('utf-8')]}
            dsp_modlist = ldap.modlist.modifyModlist(old_entry, new_entry)

            print('dsp mod list:', dsp_modlist)
            print('cn mod list:', rsc_modlist)

            if(commit):
               ldap_py.modify_s(dn, dsp_modlist)
               ldap_py.modify_s(dn, rsc_modlist)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

'''
# update_jdbc_urls()
'''
def update_jdbc_urls():

   print("update_jdbc_urls: enter")

   ldap_py = ldap.initialize('ldap://localhost:389')
   ldap_py.simple_bind_s("cn=root,cn=client,ou=novusaws,o=westgroup.com", "Ccha43mPpa.daigDINa85")
#   ldap_py.simple_bind_s("cn=root,cn=qc,ou=NOVUSAWS,o=WESTGROUP.COM", "$profesSionajj47")
#   ldap_py.simple_bind_s("cn=root,cn=qc,ou=NOVUSAWS,o=WESTGROUP.COM", "$comPP3tellymanGGdoF")
#   ldap_py.simple_bind_s("cn=root,cn=dev,ou=novusaws,o=westgroup.com", "&comPP3tellymanGGdoF")
#   ldap_py.simple_bind_s("cn=novusqc,cn=users,ou=novusaws,o=westgroup.com", "$profesSionajj47")
   print("ldaputil: ldap bound")

   baseDN = "cn=Resources,cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgDataStore)"

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:

         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0
      print()
      
      find_str = "amazonaws.com:5432//pg"
      rplc_str = "amazonaws.com:5432/pg"
      print('length:', len(result_set))

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
   
         #print('dn:', dn)
         #print(type(entry_dict), 'entry:', entry)

         jdbc_url = entry_dict.get('wgJDBCConnectionURL')

         if(jdbc_url is not None):
            jdbc_url = jdbc_url[0].decode('utf-8')
            #print(type(jdbc_url), jdbc_url)
         
            if jdbc_url.find(find_str) >= 0:
               
               count += 1
               new_jdbc_url = jdbc_url.replace(find_str, rplc_str)

                # Convert place-holders for modify-operation using modlist-module
               ldif = modlist.modifyModlist({'wgJDBCConnectionURL': entry_dict['wgJDBCConnectionURL'][0]},{'wgJDBCConnectionURL': [new_jdbc_url.encode()]})
               print()
               print(dn)
               print(ldif)
               # Do the actual modification 
               ldap_py.modify_s(dn,ldif)
      
      print()
      print(find_str)
      print("found:", count)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

   print("update_jdbc_urls: exit")


'''
# remove_ro_db_servers()
'''
def update_ro_datastores(commit):

   baseDN = "cn=Resources,cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgDataStore)"

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      # a207947-authority-02b-prod-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com
      # a207947-authority-02b-prod-pg-cluster-us-east-1.cluster-ro-ccd4069y2ydl.us-east-1.rds.amazonaws.com
      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
         
         print(entry_dict)
         print()

         cn = entry_dict.get('cn')[0].decode('utf-8')
         conn_url = entry_dict.get('wgJDBCConnectionURL')[0].decode('utf-8')

         if cn.find('docloc.ha') >= 0:
            print('catch docloc ha') 
         
         if cn.find('.update') >= 0 and conn_url.find('-ro-') < 0:
            print('found')
            print('dn:', dn)
            print('\t', conn_url)

            new_conn_url = conn_url.replace('east-1.cluster', 'east-1.cluster-ro')

            old_entry = {"wgJDBCConnectionURL": [conn_url.encode('utf-8')]}
            new_entry = {"wgJDBCConnectionURL": [new_conn_url.encode('utf-8')]}
            datastore_modlist = ldap.modlist.modifyModlist(old_entry, new_entry)
            
            print()
            print('Updating:', dn)
            print('\t', datastore_modlist)

            if commit:
               ldap_py.modify_s(dn, datastore_modlist)

   except(ldap.LDAPError) as e:
      print(e)

'''
# add_buckets
'''
def add_buckets(main_rsc, start_bkt, total_bkts, commit):

   dn = main_rsc + ",cn=Resources,cn=client,ou=novusaws,o=westgroup.com"
   cn = main_rsc

   pos = cn.find('.')
   service = cn[:pos]

   print()
   print('SERVICE:', service)
   print('\tDN:', dn)
   print('\tCN', cn)

   pos = cn.rfind('.')
   ncn = cn[:pos]

   ndn_p = 'cn='
   pos = dn.find(',')
   ndn_s = ',' + dn[pos + 1:]
   
   i = start_bkt

   for i in range(total_bkts):

      bucket = cn + str(i + start_bkt)
      ndn = 'cn=' + bucket + ',cn=' + cn + ndn_s

      print()
      print('\tncn:', ncn)
      print('\tndn:', ndn)
      print('\tbucket:', bucket)

      # A dict to help build the "body" of the object
      attrs = {}
      attrs['objectclass'] = [b'top',b'wgResource']
      attrs['cn'] = bucket.encode()
      attrs['cn'] = bucket.encode()

      # Convert our dict to nice syntax for the add-function using modlist-module
      ldif = modlist.addModlist(attrs)

      # Convert our dict to nice syntax for the add-function using modlist-module
      #ldif = modlist.addModlist(add_list)
      print("Adding: ")
      print(ldif)

      if(commit):

         try:
            ldap_py.add_s(ndn, ldif)
         except Exception as e:
            print()
            print(e)

'''
# update_bucketed_datastore_ptrs(ldap_py, log, commit)
'''
def update_bucketed_datastore_ptrs(ldap_py, log, commit):
   print('update_bucketed_datastore_ptrs(): BEGIN')
   print('update_bucketed_datastore_ptrs(): BEGIN', file=log)

   out_log = open('client_upd_dsp_0120.log', 'w')
   commit = False
   
   #baseDN = "cn=dev,ou=novusaws,o=westgroup.com"
   #baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   #resource_dict = util.get_resource_dict()

   try:
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0

      for entry in result_set:
         count = count + 1

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')

         #resource = 'nims.ccscaleretr.02'
         #rsc_loc = 'perfcsloc'
         #rsc_loc1 = 'dlc'
         
         if dsp.find('cn=resources') >= 0:
            print(rsc, dn, dsp)

         aws_rsc = None

         #if rsc in QC_RSC_DICT:
         #   aws_rsc = QC_RSC_DICT[rsc]

         if rsc in migmaps.CLIENT_BKT_RSC_DICT:

            aws_rsc = migmaps.CLIENT_BKT_RSC_DICT[rsc]

            print('    dsp:', dsp, file=log)
            print('     cn:', cn, file=log)
            print('aws_rsc:', aws_rsc, file=log)
            print('    rsc:', rsc, file=log)
            print(" ", file=log)

            if cn.find(rsc) >=0:
            
               rsc_typ = cn[cn.find('.'):]

               if re.search(r".[0-9][0-9]", rsc):
                  bucket = bucket + rsc[len(rsc)-2:]
               elif re.search(r".[0-9]", rsc): 
                  bucket = bucket + rsc[len(rsc)-1:]
               else: 
                  continue
               
               print(bucket)

               if 'master.u' in rsc_typ:
                  new_dsp = 'cn=' + rsc + 'a.update,cn=' + bucket + ',cn=' + rsc + ',cn=Resources'
                  aws_rsc = bucket
                  aws_cn = rsc + 'a.update'

               elif 'masterA.u' in rsc_typ:
                  new_dsp = 'cn=' + rsc + 'a.update,cn=' + bucket + ',cn=' + rsc+ ',cn=Resources'
                  aws_rsc = bucket
                  aws_cn = rsc + 'a.update'

               elif 'master.r' in rsc_typ:
                  new_dsp = 'cn=' + rsc + 'a.read,cn=' + bucket + ',cn=' + rsc + ',cn=Resources'
                  aws_rsc = bucket
                  aws_cn = rsc + 'a.read'

               elif 'ret-masterA.r' in rsc_typ:
                  new_dsp = 'cn=' + rsc + 'a.read,cn=' + bucket + ',cn=' + rsc + ',cn=Resources'
                  aws_rsc = bucket
                  aws_cn = rsc + 'a.read'
               
               elif 'slave.r' in rsc_typ:
                  new_dsp = 'cn=' + rsc + 'z.read,cn=' + bucket + ',cn=' + rsc + ',cn=Resources'
                  aws_rsc = bucket
                  aws_cn = rsc + '.read'
               else:
                  print('rsc_typ NOT FOUND:', rsc_typ)
            
            else:
               print(cn)
               continue
            
            #new_dsp = 'cn=' + dsp_leaf + ',cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'

            dsp_ldif = modlist.modifyModlist({'wgDataStorePtr': dsp.encode('utf-8')}, {'wgDataStorePtr': [new_dsp.encode('utf-8')]})
            
            print(dsp_ldif)
            print(dn, file=out_log)
            print('\t', dsp_ldif, file=out_log)

            if(commit):
               ldap_py.modify_s(dn, dsp_ldif)

            res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(aws_rsc).encode('utf-8')]})
            print(res_ldif)
            print('\t', res_ldif, file=out_log)

            if(commit):
               ldap_py.modify_s(dn, res_ldif)
            '''
            cn_ldif = modlist.modifyModlist({'cn': cn.encode('utf-8')}, {'cn': [(aws_cn).encode('utf-8')]})
            print(cn_ldif)
            print('\t', cn_ldif, file=out_log)
            ldap_py.modify_s(dn, cn_ldif)
            '''
         else:
            # print('\taws_rsc:', aws_rsc)
            continue
      else:
         'place holder'
         #print(rsc)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('update_bucketed_datastore_ptrs(): END')
      print('update_bucketed_datastore_ptrs(): END', file=log)

'''
# add_resource_entries(ldap_py, buckets, commit)
'''
def add_resource_entries(ldap_py, commit):
    
    secrets_client = get_client(REGION_NAME, SECRETS_MANAGER_SERVICE)

    secrets = get_secrets_list(secrets_client)

    for secret in secrets:
       secret_name = secret['Name']

       print(secret_name)

       if "a207947-novus" in secret_name and ("db-read-prod-" in secret_name or "db-update-prod-" in secret_name) and 'docloc' in secret_name:
          print('secret_name:', secret_name)
          add_resource_entry(ldap_py, secret_name, commit)

'''
###
##
# add_database_server_entry(ldap, )
##
'''
def add_database_server_entry(ldap_py, address, site, status, commit):

   # The dn of our new entry/object
   dn_suffix = ",cn=DatabaseServers,cn=client,ou=novusaws,o=westgroup.com" 
   dn = "cn=" + address + dn_suffix

   modify_list = {
      "objectclass": [b"top",b"wgDatabaseServer"],
      "cn": [address.encode()],
      "wgSite": [site.encode()],
      "wgStatus": [status.encode()]
   }

   # Convert our dict to nice syntax for the add-function using modlist-module
   ldif = modlist.addModlist(modify_list)
   print("Adding: ") 
   print('dn:', dn)
   print(ldif)

   # Do the actual synchronous add-operation to the ldapserver
   if commit:
      try:
         ldap_py.add_s(dn,ldif)
      except(ldap.LDAPError) as e:
         print(e)
      finally:
         print('EXIT')


## END add_database_server_entry()


'''
# update_jdbc_urls()
'''
def audit_jdbc_urls(ldap_py, log):

   print("audit_jdbc_urls: enter")

   baseDN = "cn=Resources,cn=prod,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   retrieveAttributes = None ## retrieve all attributes
   searchFilter = "(objectClass=wgDataStore)"

   try:
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:

         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      print('length:', len(result_set))

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
   
         #print('dn:', dn)
         #print(type(entry_dict), 'entry:', entry)

         jdbc_url = entry_dict.get('wgJDBCConnectionURL')

         juset = set()

         if(jdbc_url is not None):
            jdbc_url = jdbc_url[0].decode('utf-8')

            if 'prod' in jdbc_url:
               #if jdbc_url not in julist:
               #   julist.append(jdbc_url)

               juset.add(jdbc_url)
            
            for url in juset:
               print(url)   
               print(url, file=log)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('audit_jdbc_urls: enter')

'''
##
# add_database_servers
##
'''
def add_database_servers(ldap_py):
   
   db_clusters = util.get_db_clusters('us-east-1')
   print(type(db_clusters))

   for cluster in db_clusters:
      #print(type(cluster))
      #print('DB CLUSTER:', cluster)

      endpoint = cluster.get('Endpoint')
      reader_endpoint = cluster.get('ReaderEndpoint')
      print(endpoint)
      if '-qa-pg-cluster-' in endpoint and 'z-' not in endpoint:
         cluster_name = None
         print()
         tag_list = cluster.get('TagList')
         #print(type(tag_list), tag_list)

         for tag in tag_list:
            #print(tag['Key'], tag['Value'])
            if('tr:db-cluster' in tag['Key']):
               cluster_name =  tag['Value']
            
         if cluster_name is not None:
            
            site = (cluster_name[-1:].upper())
            if 'A' not in site and 'B' not in site:
               site = 'A'

            print('\tcluster_name', cluster_name, 'site', site)
            
            add_database_server_entry(ldap_py, endpoint, site, 'started', True)


'''
# get_xl_mig_list(workbook_name, sheet_name):
'''
def get_xl_db_list(file_name, sheet_name):

   wb = openpyxl.load_workbook(file_name)
   sheet_names = wb.sheetnames

   if sheet_name in sheet_names:
      ws = wb[sheet_name]
   else:
      print(sheet_name, 'not found. Exiting')
      return None

   max_col = ws.max_column
   max_row = ws.max_row

   print(max_col)
   print(max_row)
   print()

   migration_dict = {}
   migration_list = []
   
   for i in range(2, max_row):

      migration_dict = {
         "db_name": ws.cell(row = i, column = 1).value,
         "resource": ws.cell(row = i, column = 2).value,
         "buckets": ws.cell(row = i, column = 3).value
      }

      #print()

      migration_list.append(migration_dict)

      #for i in migration_dict:
      #   print(i, ':', migration_dict.get(i))

   return migration_list

'''
###
# add_resource_entry(ldap_py, secret, commit)
##
#  THIS IS A F'ING MESS
'''
def add_resource_entry(ldap_py, secret_name, commit):

   aws_region_name = 'us-east-1'
   aws_service_name = 'secretsmanager'
   datastore_type = 'POSTGRES NO_VARCHARS STREAMS STATEMENTCACHESIZE=20'
   jdbc_driver_name = 'org.postgresql.Driver'
   aws_region = 'us-east-1'
   read_only = 'TRUE'
   resource_dn = 'cn=Resources,cn=prod,ou=novusaws,o=westgroup.com'
   bucket_num = int(buckets)

   infra_service = ['cci', 'persist', 'util', 'loadqueue', 'loadque', 'status', 'ecpqueue', 'hadlc', 'lager', 'dlc']

   # internal check, not necessarily needed
   if 'prod' in secret_name and 'ecpqueue' in secret_name:

      session = boto3.session.Session()
      client = session.client(service_name=aws_service_name, region_name=aws_region_name)

      secret = client.get_secret_value(SecretId=secret_name)
      secret_string = secret.get('SecretString')
      secret_dict = json.loads(secret_string)

      db_host = secret_dict['host']
      db_name = secret_dict['dbname']
      user_name = secret_dict['username']

      resource_suffix = db_name[(len(db_name)-4):(len(db_name)-2)]
      sub_resource_suffix = db_name[(len(db_name)-4):(len(db_name)-1)]
      site = db_name[len(db_name)-2:len(db_name)-1]

      print()
      print('site:', site)
      print('resource_suffix:', resource_suffix)
      print('sub_resource_suffix:', sub_resource_suffix)

      if secret_name.find('update') >= 0:
         base_resource_name = secret_name[14:-29]
         base_resource_name = base_resource_name.replace('-', '.')
         print('base_resource_name:', base_resource_name)
      else:
         base_resource_name = secret_name[14:-27]
         base_resource_name = base_resource_name.replace('-', '.')
         print('base_resource_name:', base_resource_name)
      
      pos = base_resource_name.find('.')
      
      if(pos >= 0):
         service_name = base_resource_name[:pos]
      else:
         service_name = base_resource_name

      print('service_name:', service_name)

      if(service_name in infra_service):
         resource_name = base_resource_name
         sub_resource_name = service_name + '.' + sub_resource_suffix
         ##########!!!!!!!!!!!!!!!!!!TEMP SKIP!!!!!!!!!!!!!!!!!!!!!################
         #return
      else:
         resource_name = base_resource_name + '.' + resource_suffix
         sub_resource_name = base_resource_name + '.' + sub_resource_suffix

      if secret_name.find('update') >= 0:
         read_only = 'FALSE'
         sub_resource_name = sub_resource_name + '.update'
         jdbc_connection_url = 'jdbc:postgresql://' + db_host + ':5432/' + db_name + '?sslmode=verify-ca&sslrootcert=/novus/prod/releases/db/rds-ca-2019-root.pem'
         print("sub_resource_name: " + sub_resource_name)
      else:
         read_only = 'TRUE'
         read_db_host = db_host.replace('-east-1.cluster-cc', '-east-1.cluster-ro-cc')
         sub_resource_name = sub_resource_name + '.read'
         jdbc_connection_url = 'jdbc:postgresql://' + read_db_host + ':5432/' + db_name + '?sslmode=verify-ca&sslrootcert=/novus/prod/releases/db/rds-ca-2019-root.pem'
         print("sub_resource_name: " + sub_resource_name)

      resource_dn = 'cn=' + resource_name + ',cn=Resources,cn=prod,ou=novusaws,o=westgroup.com'
      datastore_dn = 'cn=' + sub_resource_name + ',' + resource_dn
      
      print()
      print('db_host: ' + db_host)
      print('db_name: ' + db_name)
      print("user_name: " + user_name)
      print("service_name: " + service_name)
      print("resource_name: " + resource_name)
      print("sub_resource_name: " + sub_resource_name)
      print("jdbc_connection_url: " + jdbc_connection_url)
      print("datastore_dn: " + datastore_dn)

      bucket_num = 0
      found = False

      '''
      for db in TARGET_DBS:
         if resource_name == db['resource']:
            if db['buckets'] is not None:
               bucket_num = db['buckets']

            found = True
            break
      
      print('FOUND:', found)
      '''

      if bucket_num > 0:
         resource_modify_list = {
            "objectclass": [b"top",b"wgResource"],
            "cn": [resource_name.encode()],
            "wgCurrentBuckets": [str(bucket_num).encode()],
            "wgNewBuckets": [str(bucket_num).encode()]
         }
      else:
         resource_modify_list = {
            "objectclass": [b"top",b"wgResource"],
            "cn": [resource_name.encode()],
         }
      
      datastore_modify_list = {
         "objectclass": [b"top",b"wgDataStore"],
         "cn": [sub_resource_name.encode()],
         "wgAWSRegion": [aws_region.encode()],
         "wgAWSSecretName": [secret_name.encode()],
         "wgDatabaseHost": [db_host.encode()],
         "wgDataStoreType": [datastore_type.encode()],
         "wgJDBCConnectionURL": [jdbc_connection_url.encode()],
         "wgJDBCDriverName": [jdbc_driver_name.encode()],
         "wgReadOnly": [read_only.encode()],
         "wgResourceName": [resource_name.encode()]
      }

      resource_ldif = modlist.addModlist(resource_modify_list)
      print('Adding: ', resource_dn)
      print('        ', resource_ldif)
      
      if commit:
         try:
            ldap_py.add_s(resource_dn, resource_ldif)
         except ldap.ALREADY_EXISTS:
            print('Entry already Exists on ldap server.', resource_dn)
         except ldap.LDAPError:
            print("Failed to add entry to the ldap server", resource_dn)
      
      datastore_ldif = modlist.addModlist(datastore_modify_list)
      print('Adding: ', datastore_dn)
      print('        ', datastore_ldif)
      print()

      if commit:
         try:
            ldap_py.add_s(datastore_dn, datastore_ldif)
         except ldap.ALREADY_EXISTS:
            print('Entry already Exists on ldap server.', datastore_dn)
         except ldap.LDAPError:
            print("Failed to add entry to the ldap server", datastore_dn)

'''
# add_bucket_datastores
'''
def add_bucket_datastores(main_resource, start_bucket, buckets, db_override, commit):

   ds_attrs = {}
   ds_attrs['objectclass'] = [b'top',b'wgDataStore']
   ds_attrs['wgAWSRegion'] = [b'us-east-1']
   ds_attrs['wgDataStoreType'] = [b'POSTGRES NO_VARCHARS STREAMS STATEMENTCACHESIZE=20']
   ds_attrs['wgJDBCDriverName'] = [b'org.postgresql.Driver']

   update_ar = True
   update_au = True
      
   for i in range(buckets):

      bucket_num = str(i + start_bucket)
      bucket_resource = main_resource + bucket_num

      if(update_ar):
         '''
            primary read datastore
         '''
         if db_override is not None:
            cn = db_override.replace('-', '.') + 'a.read'
         else:
            cn = main_resource + 'a.read'

         # allow override for non-conventional names. usually buckets split across resource names.
         if db_override is None:
            db_name = main_resource.replace('.', '-')
         else:
            db_name = db_override
         
         pri_db_name = db_name + 'a'

         dn = 'cn=' + cn + ',cn='
         dn = dn + main_resource + bucket_num + ',cn='
         dn = dn + main_resource + ',cn=Resources,cn=prod,ou=novusaws,o=westgroup.com'

         cluster_name = db_name.replace('-', '')
         cluster_name = 'pg' + cluster_name + 'ap'
         
         secret_name = 'a207947-novus-' + pri_db_name + '-db-read-prod-us-east-1'
         db_host = 'a207947-' + pri_db_name + '-prod-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com'
         jdbc_url = 'jdbc:postgresql://a207947-' + pri_db_name + '-prod-pg-cluster-us-east-1.cluster-ro-ccd4069y2ydl.us-east-1.rds.amazonaws.com:5432//' + cluster_name + '?sslmode=verify-ca&sslrootcert=/novus/prod/releases/db/rds-ca-2019-root.pem'
         
         ds_attrs['cn'] = [cn.encode()]
         ds_attrs['wgAWSSecretName'] = [secret_name.encode()]
         ds_attrs['wgDatabaseHost'] = [db_host.encode()]
         ds_attrs['wgJDBCConnectionURL'] = [jdbc_url.encode()]
         ds_attrs['wgReadOnly'] = [b'TRUE']
         ds_attrs['wgResourceName'] = [bucket_resource.encode()]
         
         ldif = modlist.addModlist(ds_attrs)
         #ldif = modlist.modifyModlist(ds_attrs, ds_attrs)
         print()
         print("Adding: ", dn)
         print(ldif)
         
         if commit:
         
            try:
               ldap_py.add_s(dn, ldif)
            except ldap.ALREADY_EXISTS:
               print('Entry already Exists on ldap server.', dn)
            except ldap.LDAPError:
               print("Failed to add entry to the ldap server", dn)

      if(update_au):

         if db_override is not None:
            cn = db_override.replace('-', '.') + 'a.update'
         else:
            cn = main_resource + 'a.update'

         # allow override for non-conventional names. usually buckets split across resource names.
         if db_override is None:
            db_name = main_resource.replace('.', '-')
         else:
            db_name = db_override

         pri_db_name = db_name + 'a'

         dn = 'cn=' + cn + ',cn='
         dn = dn + main_resource + bucket_num + ',cn='
         dn = dn + main_resource + ',cn=Resources,cn=prod,ou=novusaws,o=westgroup.com'

         cluster_name = db_name.replace('-', '')
         cluster_name = 'pg' + cluster_name + 'ap'
         
         secret_name = 'a207947-novus-' + pri_db_name + '-db-update-prod-us-east-1'
         db_host = 'a207947-' + pri_db_name + '-prod-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com'
         jdbc_url = 'jdbc:postgresql://a207947-' + pri_db_name + '-prod-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com:5432//' + cluster_name + '?sslmode=verify-ca&sslrootcert=/novus/prod/releases/db/rds-ca-2019-root.pem'
         
         ds_attrs['cn'] = [cn.encode()]
         ds_attrs['wgAWSSecretName'] = [secret_name.encode()]
         ds_attrs['wgDatabaseHost'] = [db_host.encode()]
         ds_attrs['wgJDBCConnectionURL'] = [jdbc_url.encode()]
         ds_attrs['wgReadOnly'] = [b'FALSE']
         ds_attrs['wgResourceName'] = [bucket_resource.encode()]
         
         ldif = modlist.addModlist(ds_attrs)
         print()
         print("Adding: ", dn)
         print(ldif)


         
         if commit:

            try:
               ldap_py.add_s(dn, ldif)
            except ldap.ALREADY_EXISTS:
               print('Entry already Exists on ldap server.', dn)
            except ldap.LDAPError:
               print("Failed to add entry to the ldap server", dn)
'''
# add_bucket_resources        
'''
def add_bucket_resources(main_resource, start_bucket, buckets, commit):

   dn = 'cn=' + main_resource + ',cn=Resources,cn=prod,ou=novusaws,o=westgroup.com'
   ## DEBUG OVERRIDES

   for bucket in range(buckets):

      current_bucket = bucket + start_bucket 
      # A dict to help build the "body" of the object
      attrs = {}
      attrs['objectclass'] = [b'top',b'wgResource']
      attrs['cn'] = (main_resource + str(current_bucket)).encode()

      # Convert our dict to nice syntax for the add-function using modlist-module
      ldif = modlist.addModlist(attrs)

      #mod_attrs = [( ldap.MOD_ADD, 'cn', (bucket_resource + '.bucket.' + str(current_bucket)).encode())]
      add_dn = 'cn=' + main_resource + str(current_bucket) + ',' + dn 

      print()
      print('add_dn:', add_dn)
      print('ldif', ldif)
      #print('mod_attrs', mod_attrs)

      if(commit):

         try:
            ldap_py.add_s(add_dn, ldif)
            #ldap_py.modify_s(add_dn,mod_attrs)
         except Exception as e:
            print()
            print(e)

'''
# update_datastore_ptrs(ldap_py, log, commit)
'''
def fix_dlc_oneoff(ldap_py, commit):

   baseDN = "cn=prod,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   retrieveAttributes = None ## retrieve all attributes
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   try:
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')

         if 'dlc' in rsc and 'hadlc' not in rsc:
            print('rsc:', rsc)

            new_rsc = rsc[:3]

            print('new_rsc:', new_rsc)

            res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(new_rsc).encode('utf-8')]})
            print('\t',res_ldif)

            if(commit):
               ldap_py.modify_s(dn, res_ldif)
   
   except(ldap.LDAPError) as e:
      print(e)     

'''
# update_datastores()
'''
def update_datastores(ldap_py, log, commit):
   print('update_datastores(): BEGIN')
   print('update_datastores(): BEGIN', file=log)

   baseDN = "cn=Resources,cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None 
   searchFilter = "(objectClass=wgDataStore)"

   try: 
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count1 = 0

      print()

      for entry in result_set:
         count1 = count1 + 1

         dn = entry[0][0]
         entry_dict = entry[0][1]

         if entry_dict is not None:
            cn = entry_dict.get('cn')[0].decode('utf-8')
            secret_name = entry_dict.get('wgAWSSecretName')[0].decode('utf-8')
            read_only = entry_dict.get('wgReadOnly')[0].decode('utf-8')

            if('update' in secret_name and 'TRUE' in read_only):
               #print()
               print(dn)
               print('\t',secret_name)
               #print(read_only)
            

               ldif = modlist.modifyModlist({'wgReadOnly': entry_dict['wgReadOnly'][0]},{'wgReadOnly': ["FALSE".encode('utf-8')]})
               print(ldif) 

               if(commit):
                  ldap_py.modify_s(dn,ldif)
         
         

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')


'''
# update_datastore_ptrs_oneoff(ldap_py, log, commit)
'''
def update_datastore_ptrs_oneoff(ldap_py, log, commit):
   print('update_datastore_ptrs_oneoff(): BEGIN')
   print('update_datastore_ptrs_oneoff(): BEGIN', file=log)

   out_log = open('client_upd_dsp_0122.log', 'w')
   print('output log:', out_log)

   #commit = False
   count = 0
   
   #baseDN = "cn=dev,ou=novusaws,o=westgroup.com"
   #baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   baseDN = "cn=client,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   ## retrieve all attributes - again adjust to your needs - see documentation for more options
   retrieveAttributes = None
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   #resource_dict = util.get_resource_dict()

   try:
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0
      op_count = 0

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')
         
         rsc = rsc.lower()
         #cn = cn.lower()

         #if dsp.find('cn=resources') or dsp.find('cn=Resources'):

         #"wwlkccitingrefss":  "norm.intl.01",

         if 'normwest1' in cn:

            #print(rsc, dn, dsp)
            op_count = op_count + 1

            #if dsp.find('wkrnims') >= 0:
            #   print('BREAK')

            aws_rsc = 'norm.shared.01'
            dsp_split = dsp.split('cn=') 
            base_rsc =  dsp_split[len(dsp_split) - 2]
            base_rsc = base_rsc[:len(base_rsc) - 1]
            base_rsc = base_rsc.lower()

            infra_rsc = False

            print('    dsp:', dsp, file=log)
            print('     cn:', cn, file=log)
            print('aws_rsc:', aws_rsc, file=log)
            print('    rsc:', rsc, file=log)
            print(" ", file=log)

            print('    dsp:', dsp)
            print('     cn:', cn)
            print('aws_rsc:', aws_rsc)
            print('    rsc:', rsc)
            print(" ")

            rsc_typ = cn[cn.find('.'):]

            aws_rsc_b = aws_rsc
            bucket = aws_rsc + rsc[len(base_rsc):]

            is_bucket = True
            skip = False


            if is_bucket and not skip:
               if 'master.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.update,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket

               elif 'masterA.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.update,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket

               elif '.Site-A.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.update,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket

               elif '.Site-A.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket

               elif '.Site-B.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'b.update,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket
               
               elif '.Site-B.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'b.read,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket

               elif 'master.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket
                  
               elif 'ret-masterA.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket
               elif 'slave.slave.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket
                  aws_cn = rsc + 's.read'
               elif 'slave.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc + ',cn=Resources'
                  aws_rsc = bucket
                  aws_cn = rsc + 's.read'
               
               else:
                  print('rsc_typ NOT FOUND:', rsc_typ)
                  continue

            elif not skip:
               if 'master.u' in rsc_typ:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'

                  aws_rsc = bucket

               elif 'master.1.u' in rsc_typ and infra_rsc:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket
               
               elif 'master.2.u' in rsc_typ:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'b.update,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket

               elif 'masterA.u' in rsc_typ:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'

                  aws_rsc = bucket

               elif '.Site-A.u' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'

                  aws_rsc = bucket

               elif '.Site-A.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket

               elif '.Site-B.u' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'b.update,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket
               
               elif '.Site-B.r' in rsc_typ:
               
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'b.read,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket

               elif 'master.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket
                  
               elif 'ret-masterA.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket
               
               elif 'slave.r' in rsc_typ:
                  
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'z.read,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket
               
               elif 'slave.u' in rsc_typ:
                  
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'z.read,cn=' + aws_rsc + ',cn=Resources'
                  
                  aws_rsc = bucket

               else:
                  print('rsc_typ NOT FOUND:', rsc_typ)
                  continue   
            else:
               print('SKIP DOUBLE SLAVE OR DLC')

            if not skip:
               dsp_ldif = modlist.modifyModlist({'wgDataStorePtr': dsp.encode('utf-8')}, {'wgDataStorePtr': [new_dsp.encode('utf-8')]})

               print('\t', dn)
               print('\t\t', dsp_ldif)
               print(dn, file=out_log)
               print('\t\t', dsp_ldif, file=out_log)

               if(commit):
                  ldap_py.modify_s(dn, dsp_ldif)
               
               count = count + 1

               res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(aws_rsc).encode('utf-8')]})
               print('\t\t',res_ldif)
               print('\t\t', res_ldif, file=out_log)

               if(commit):
                  ldap_py.modify_s(dn, res_ldif)
            
               '''
               cn_ldif = modlist.modifyModlist({'cn': cn.encode('utf-8')}, {'cn': [(aws_cn).encode('utf-8')]})
               print(cn_ldif)
               print('\t', cn_ldif, file=out_log)
               
               if commit:
                  ldap_py.modify_s(dn, cn_ldif)
            '''
         else:
            ph = 'place holder'

            #if dn.find('wwlkc') >= 0:
            #   print('skipping rsc, \'resources\' not matched:', rsc,'dn:', dn, 'dsp:', dsp)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('update_datastore_ptrs_oneoff(): END')
      print('update_datastore_ptrs_oneoff(): END', file=log)

'''
# update_datastore_stub()
'''
def update_datastore_stub(ldap_py, log, commit):
   print('update_datastore_stub(): BEGIN')
   print('update_datastore_stub(): BEGIN', file=log)

   #baseDN = "cn=Resources,cn=prod,ou=novusaws,o=westgroup.com"
   baseDN = "cn=prod,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   retrieveAttributes = None # actually means retrieve all, add values to limit
   #searchFilter = "(&(objectClass=wgDataStore)(cn=nort.trtax.*))"
   #searchFilter = "(objectClass=wgDataStore)"
   searchFilter = "(objectClass=wgAccessibleDatastore)"
   #searchFilter = "(&(objectClass=wgDataStore)(wgResourceName=*cn=*))"

   try: 
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      for entry in result_set:

         dn = entry[0][0]
         
         if 'normwest1' in dn.lower():
            dn_list = dn.split(',')
            
            print()
            print(dn_list)
            print(entry)

         entry_dict = entry[0][1]


         '''
         {'wgAWSSecretName': [b'a207947-novus-cci-01a-db-read-prod-us-east-1'], 
          'wgReadOnly': [b'TRUE'],
          'wgDataStoreType': [b'POSTGRES NO_VARCHARS STREAMS STATEMENTCACHESIZE=20'], 
          'wgJDBCConnectionURL': [b'jdbc:postgresql://a207947-cci-01a-prod-pg-cluster-us-east-1.cluster-ro-ccd4069y2ydl.us-east-1.rds.amazonaws.com:5432/pgcci01ap?sslmode=verify-ca&sslrootcert=/novus/prod/releases/db/rds-ca-2019-root.pem'], 
          'wgJDBCDriverName': [b'org.postgresql.Driver'], 
          'cn': [b'cci.01a.read'], 
          'wgAWSRegion': [b'us-east-1'], 
          'wgResourceName': [b'cci'], 
          'wgDatabaseHost': [b'a207947-cci-01a-prod-pg-cluster-us-east-1.cluster-ccd4069y2ydl.us-east-1.rds.amazonaws.com'], 
          'objectClass': [b'top', b'wgDataStore']}
         '''

         if entry_dict is not None:

            if(False):
               print(entry_dict)

            '''
               new_rsc_name = dn_list[1][3:] # strip 'cn='

               print()
               print(new_rsc_name)

               ldif = modlist.modifyModlist({'wgResourceName': entry_dict['wgResourceName'][0]},{'wgResourceName': [new_rsc_name.encode('utf-8')]})
               
               print(dn)
               print(ldif) 

               if(commit):
                  ldap_py.modify_s(dn,ldif)
            '''

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

PROD = "PROD"
CLIENT = "QA"

'''
# update_datastore_ptr_stub()
'''
def update_datastore_ptr_stub(ldap_py, log, commit):
   print('update_datastore_stub(): BEGIN')
   print('update_datastore_stub(): BEGIN', file=log)

   baseDN = "cn=prod,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   retrieveAttributes = None # actually means retrieve all, add values to limit

   #searchFilter = "(&(objectClass=wgAccessibleDatastore)(wgDataStorePtr=*resources*))"
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   unmapped = []

   try: 
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')

         ## DEBUG LINE ##
         #print(cn, rsc, dsp, dn)
         
         rsc = rsc.lower()

         # override, or not, for specific DBs
         if ('resources' in dsp and ('shp.rac3.n1' in dsp.lower() or 'normshared1' in dsp.lower())):

            aws_rsc = None
            dsp_split = dsp.split('cn=') 
            base_rsc =  dsp_split[len(dsp_split) - 2]
            base_rsc = base_rsc[:len(base_rsc) - 1]
            base_rsc = base_rsc.lower()

            rsc_typ = cn[cn.find('.'):]

            
            if base_rsc in migmaps.CLIENT_INF_RSC_DICT:
               aws_rsc = migmaps.CLIENT_INF_RSC_DICT[base_rsc]
               infra_rsc = True
            elif base_rsc in migmaps.PROD_MASTER_RSC_LIST_NI:
               aws_rsc = migmaps.PROD_MASTER_RSC_LIST_NI[base_rsc]
            else:
               print('RESOURCE NOT MAPPED:', base_rsc)
               print('\t', dsp)
               
            
            if base_rsc not in unmapped:
               unmapped.append(base_rsc)
            else:
               print(base_rsc, dsp)

            '''
            if 'cn=dlc' not in dsp:
               continue
            else:
               #print(dsp.split(','))
               aws_rsc = dsp.split(',')[1][3:]
            '''

            dn_commit = dn

            '''
            print(" ", file=log)
            print('     dn:', dn, file=log)
            print('    dsp:', dsp, file=log)
            print('     cn:', cn, file=log)
            print('aws_rsc:', aws_rsc, file=log)
            print('    rsc:', rsc, file=log)
            print(" ", file=log)

            print()
            print('     dn:', dn)
            print('    dsp:', dsp)
            print('     cn:', cn)
            print('aws_rsc:', aws_rsc)
            print('    rsc:', rsc)
            print('rsc_typ:', rsc_typ)
            '''

            '''
            if "docloc.01," in dsp:
               new_dsp = dsp.replace("docloc.01,", "dlc,")
            elif "docloc.02," in dsp:
               new_dsp = dsp.replace("docloc.02,", "dlc,")
            elif "docloc.03," in dsp:
               new_dsp = dsp.replace("docloc.03,", "dlc,")
            elif "docloc.04," in dsp:
               new_dsp = dsp.replace("docloc.04,", "dlc,")
            '''


            '''
            if 'persist0,' in dsp:
               new_dsp = dsp.replace(".persista.", ".persist0.")
            elif 'persist1,' in dsp:
               new_dsp = dsp.replace(".persist0.", ".persist1.")
            elif 'persist2,' in dsp:
               new_dsp = dsp.replace(".persist0.", ".persist2.")
            elif 'persist3,' in dsp:
               new_dsp = dsp.replace(".persist0.", ".persist3.")
            '''

            '''
            res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(aws_rsc).encode('utf-8')]})
            print('DN:', dn_commit)
            #print('\t\t',res_ldif)
            #print('\t\t', res_ldif, file=log)

            if(commit):
               ldap_py.modify_s(dn_commit, res_ldif)

                        
            dsp_ldif = modlist.modifyModlist({'wgDataStorePtr': dsp.encode('utf-8')}, {'wgDataStorePtr': [new_dsp.encode('utf-8')]})

            if 'cn=dlc,cn=Resources' in new_dsp:

               print('\t', dn_commit)
               
               
               print('\t\t', dsp_ldif)
               print(dn_commit, file=log)
               print('\t\t', dsp_ldif, file=log)
               
            
               if(commit):
                  ldap_py.modify_s(dn_commit, dsp_ldif)

            
               
               res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(aws_rsc).encode('utf-8')]})
               print('\t\t',res_ldif)
               print('\t\t', res_ldif, file=log)

               if(commit):
                  ldap_py.modify_s(dn_commit, res_ldif)
            else:
               print()
               print('skipping:') 
               print('\t', dn) 
               print('\t', dsp)
               print('\t', new_dsp)
               print()
            '''
         else:
            print('.', end='')
            #print(dn)
            #print('\t', dsp)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

'''
# datastore_ptr_audit()
'''
def datastore_ptr_audit(ldap_py, log, commit):
   print('update_datastore_stub(): BEGIN')
   print('update_datastore_stub(): BEGIN', file=log)

   baseDN = "cn=prod,ou=novusaws,o=westgroup.com"
   searchScope = ldap.SCOPE_SUBTREE
   retrieveAttributes = None # actually means retrieve all, add values to limit

   #searchFilter = "(&(objectClass=wgAccessibleDatastore)(wgDataStorePtr=*resources*))"
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   unmapped = []

   try: 
      
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')

         ## DEBUG LINE ##
         #print(cn, rsc, dsp, dn)
         
         rsc = rsc.lower()

         # override, or not, for specific DBs
         #if ('resources' in dsp and ('shp.rac3.n1' in dsp.lower() or 'normshared1' in dsp.lower())):
         if 'persist' in dsp.lower() :

            aws_rsc = None
            dsp_split = dsp.split('cn=') 
            base_rsc =  dsp_split[len(dsp_split) - 2]
            base_rsc = base_rsc[:len(base_rsc) - 1]
            base_rsc = base_rsc.lower()

            rsc_typ = cn[cn.find('.'):]

            print(dn)

            if base_rsc in migmaps.CLIENT_INF_RSC_DICT:
               aws_rsc = migmaps.CLIENT_INF_RSC_DICT[base_rsc]
               infra_rsc = True
            elif base_rsc in migmaps.PROD_MASTER_RSC_LIST_NI:
               aws_rsc = migmaps.PROD_MASTER_RSC_LIST_NI[base_rsc]
            else:
               
               if(False):
                  print('RESOURCE NOT MAPPED:', base_rsc)
                  print('\t', dsp)
               
               
            
            if base_rsc not in unmapped:
               unmapped.append(base_rsc)
            else:
               if(False):
                  print(base_rsc, dsp)

            dn_commit = dn

         else:
            print('.', end='')
            #print(dn)
            #print('\t', dsp)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')

'''
# update_datastore_ptrs(ldap_py, log, commit)
'''
def update_datastore_ptrs(ldap_py, log, commit):
   print('update_datastore_ptrs(): BEGIN')
   print('update_datastore_ptrs(): BEGIN', file=log)

   count = 0
   unmapped = []
   
   baseDN = "cn=prod,ou=novusaws,o=westgroup.com"

   searchScope = ldap.SCOPE_SUBTREE
   retrieveAttributes = None ## retrieve all attributes
   #searchFilter = "(&(objectClass=wgAccessibleDatastore)(wgResourceName=*cn=*))"
   searchFilter = "(objectClass=wgAccessibleDatastore)"
   #searchFilter = "(&(objectClass=wgAccessibleDatastore)(cn=*persist*))"

   try:
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)

      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0
      op_count = 0

      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')

         ## DEBUG LINE ##
         #print(cn, rsc, dsp, dn)
         
         rsc = rsc.lower()
         hvr = False

         if 'ccireplica' in rsc:
            rsc = 'cci'
            print('KLUDGE RSC:', rsc)

         # override, or not, for specific DBs
         #if 'resources' in dsp and 'metadoc' in dsp.lower():

         aws_rsc = None
         dsp_split = dsp.split('cn=') 
         base_rsc =  dsp_split[len(dsp_split) - 2]
         base_rsc = base_rsc[:len(base_rsc) - 1]
            
         base_rsc = base_rsc.lower()
         rsc_typ = cn[cn.find('.'):]

         if 'wlnuddoc.rac' in rsc.lower():
         #if 'resources' in dsp.lower() and 'metadoc' in dsp.lower():
         #if 'nop38' in dn.lower() and 'norm.aunz' in base_rsc and 'cn=norm' in dn.lower():

            rsc = rsc[:len(rsc) - 6]

            op_count = op_count + 1

            #print()
            #print(base_rsc)

            infra_rsc = False
            
            kludge = False

            if base_rsc in migmaps.CLIENT_INF_RSC_DICT:
               aws_rsc = migmaps.CLIENT_INF_RSC_DICT[base_rsc]
               infra_rsc = True
            elif base_rsc in migmaps.PROD_MASTER_RSC_LIST:
               aws_rsc = migmaps.PROD_MASTER_RSC_LIST[base_rsc]
            elif(kludge):
               aws_rsc = 'norm.ml.citgref.01'
            else:
               print('RESOURCE NOT MAPPED:', base_rsc, dn)
               
               if base_rsc not in unmapped:
                  unmapped.append(base_rsc)
               
               continue
            
            dn_commit = dn

            if('r.ret' in dn):
               dn = dn.replace('r.ret', '.ret')
               cn = cn.replace('r.ret', '.ret')
               rsc = rsc[:-1]
               #dsp = dsp.replace('r.ret', '.ret')

            print(" ", file=log)
            print('     dn:', dn, file=log)
            print('    dsp:', dsp, file=log)
            print('     cn:', cn, file=log)
            print('aws_rsc:', aws_rsc, file=log)
            print('    rsc:', rsc, file=log)
            print(" ", file=log)

            print()
            print('     dn:', dn)
            print('    dsp:', dsp)
            print('     cn:', cn)
            print('aws_rsc:', aws_rsc)
            print('    rsc:', rsc)
            print('rsc_typ:', rsc_typ)

            aws_rsc_b = aws_rsc

            #kludge

            bucket = aws_rsc + rsc[len(base_rsc):]

            is_bucket = False
            is_dlc = False
            is_auth = False
            skip = False

            # double slave fix
            if bucket.find('slave') >= 0: # or 
               bucket = bucket.strip('.slave')
               print('bucket:', bucket)
               skip = False

            # DLC FIX
            if 'dlc' in bucket.lower() and 'hadlc' not in bucket.lower():
               bucket_num = int(bucket[3:])
               is_bucket = True
               is_dlc = True
               skip = False
               
               if bucket_num < 8:
                  aws_rsc_b = 'docloc.01'
               elif bucket_num > 7 and bucket_num < 16:
                  aws_rsc_b = 'docloc.02'
               elif bucket_num > 15 and bucket_num < 24:
                  aws_rsc_b = 'docloc.03'
               else:
                  aws_rsc_b = 'docloc.04'

               print('bucket:', bucket, 'aws_rsc:', aws_rsc)

            if 'ecpqueue' in bucket.lower():
               aws_rsc_b = 'ecpqueue.01'
               print('bucket:', bucket, 'aws_rsc:', aws_rsc)

            if 'auth' in bucket.lower():
               is_bucket = True
               is_auth = True

               if 'wlnud' in rsc.lower():
                  bucket = 'authority.01'
                  aws_rsc_b = bucket   
                  print('01 rsc_typ:', rsc_typ)

               elif 'predocket' in rsc.lower():
                  bucket = 'authority.02'
                  aws_rsc_b = bucket
                  print('02 rsc_typ:', rsc_typ)

            else:
               #print()
               if len(bucket) == len(aws_rsc):
                  print('THIS is NOT a bucket:', bucket)
                  is_bucket = False
               else:
                  print('THIS is A bucket:', bucket)
                  is_bucket = True

            if is_bucket and not skip:

               if is_auth is not True and is_dlc is not True:
                     aws_rsc = bucket

               if 'master.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.update,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'

               elif 'masterA.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.update,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
   
               elif '.Site-A.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.update,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
   
               elif '.Site-A.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
   
               elif '.Site-B.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'b.update,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
               
               elif '.Site-B.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'b.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'

               elif 'master.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
                  
               elif 'ret-masterA.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'

               elif 'ret-Site-A.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'

               elif 'ret-Site-B.r' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'b.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'

               elif 'slave.slave.r' in rsc_typ:
                  if hvr:
                     new_dsp = 'cn=' + aws_rsc_b + 'b.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
                  else:
                     new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
               
               elif 'slave.r' in rsc_typ:
                  if hvr:
                     new_dsp = 'cn=' + aws_rsc_b + 'b.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
                  else:
                     new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
               
               elif rsc_typ =='.slave':
                  if hvr:
                     new_dsp = 'cn=' + aws_rsc_b + 'b.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
                  else:
                     new_dsp = 'cn=' + aws_rsc_b + 'a.read,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'

               elif '.load-Site-A.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'a.update,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'
               
               elif '.load-Site-B.u' in rsc_typ:
                  new_dsp = 'cn=' + aws_rsc_b + 'b.update,cn=' + bucket + ',cn=' + aws_rsc_b + ',cn=Resources'

               else:
                  print('*1*')
                  print('\tdn', dn)
                  print('\t\trsc_typ NOT FOUND:', rsc_typ)
                  print('*1*')

                  continue

            elif not skip:
               
               aws_rsc = bucket

               if 'master.u' in rsc_typ:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'

               elif 'master.1.u' in rsc_typ:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'
               
               elif 'master.2.u' in rsc_typ:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'b.update,cn=' + aws_rsc + ',cn=Resources'

               elif 'masterA.u' in rsc_typ:
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'

               elif '.Site-A.u' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'

               elif '.Site-A.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
   
               elif '.Site-B.u' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'
               
               elif '.Site-B.r' in rsc_typ:
               
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'

               elif 'master.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
                  
               elif 'ret-masterA.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'

               elif 'ret-Site-B.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif 'ret-Site-A.r' in rsc_typ:

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif rsc_typ == '.slave':

                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif 'slave.r' in rsc_typ:
                  
                  if infra_rsc:
                     if hvr:
                        new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'   
                     else:
                        new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif 'slave.1.r' in rsc_typ:
                  
                  if infra_rsc:
                     if hvr:
                        new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'
                     else:
                        new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif 'slave.2.r' in rsc_typ:
                  
                  if infra_rsc:
                     if hvr:
                        new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'
                     else:
                        new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'

               elif 'master.1.r' in rsc_typ:
                  
                  if infra_rsc:
                     if hvr:
                        new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'
                     else:
                        new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif 'master.2.r' in rsc_typ:
                  
                  if infra_rsc:
                     if hvr:
                        new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'
                     else:
                        new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif 'slave.u' in rsc_typ:
                  
                  if infra_rsc:
                     if hvr:
                        new_dsp = 'cn=' + aws_rsc + '.01b.update,cn=' + aws_rsc + ',cn=Resources'   
                     else:
                        new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'
               
               elif '.01a.read' in rsc_typ:
                  
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.read,cn=' + aws_rsc + ',cn=Resources'

               elif '.01b.read' in rsc_typ:
                  
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.read,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'b.read,cn=' + aws_rsc + ',cn=Resources'

               elif '.01a.update' in rsc_typ:
                  
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01a.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'a.update,cn=' + aws_rsc + ',cn=Resources'

               elif '.01b.update' in rsc_typ:
                  
                  if infra_rsc:
                     new_dsp = 'cn=' + aws_rsc + '.01b.update,cn=' + aws_rsc + ',cn=Resources'   
                  else: 
                     new_dsp = 'cn=' + aws_rsc + 'b.update,cn=' + aws_rsc + ',cn=Resources'

               else:
                  print('*2*')
                  print('\tdn', dn)
                  print('\t\trsc_typ NOT FOUND:', rsc_typ)
                  print('*2*')

                  continue   
            else:
               print('SKIP DOUBLE SLAVE OR DLC')

            if not skip:
               dsp_ldif = modlist.modifyModlist({'wgDataStorePtr': dsp.encode('utf-8')}, {'wgDataStorePtr': [new_dsp.encode('utf-8')]})

               print('\t', dn_commit)
               print('\t\t', dsp_ldif)
               print(dn_commit, file=log)
               print('\t\t', dsp_ldif, file=log)

               if(commit):
                  ldap_py.modify_s(dn_commit, dsp_ldif)
               
               count = count + 1

               res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(aws_rsc).encode('utf-8')]})
               print('\t\t',res_ldif)
               print('\t\t', res_ldif, file=log)

               if(commit):
                  ldap_py.modify_s(dn_commit, res_ldif)
            
               '''
               cn_ldif = modlist.modifyModlist({'cn': cn.encode('utf-8')}, {'cn': [(aws_cn).encode('utf-8')]})
               print(cn_ldif)
               print('\t', cn_ldif, file=out_log)
               
               if commit:
                  ldap_py.modify_s(dn, cn_ldif)
               '''
         else:
            print('.', end='')
            #print("SKIPPING RESOURCE:", rsc, cn, dsp, dn)
            
      else:
         ph = 'place holder'
         #print()
         #print('SKIPPING RESOURCE, \'resources\' not matched:', rsc,'dn:', dn, 'dsp:', dsp)

      #for rsc in unmapped:
      #   print(rsc)

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('update_datastore_ptrs(): END')
      print('update_datastore_ptrs(): END', file=log)

      
'''
# search_datastore_ptrs(ldap_py, log, commit)
'''
def search_datastore_ptrs(ldap_py):

   baseDN = "cn=prod,ou=novusaws,o=westgroup.com"

   searchScope = ldap.SCOPE_SUBTREE
   retrieveAttributes = None ## retrieve all attributes
   searchFilter = "(objectClass=wgAccessibleDatastore)"

   try:
      ldap_result_id = ldap_py.search(baseDN, searchScope, searchFilter, retrieveAttributes)
      result_set = []

      while 1:
         
         result_type, result_data = ldap_py.result(ldap_result_id, 0)

         if (result_data == []):
            break
         else:
            if result_type == ldap.RES_SEARCH_ENTRY:
               result_set.append(result_data)

      count = 0
      op_count = 0

      print()
      for entry in result_set:

         dn = entry[0][0]
         entry_dict = entry[0][1]
         cn = entry_dict.get('cn')[0].decode('utf-8')
         dsp = entry_dict.get('wgDataStorePtr')[0].decode('utf-8')
         rsc = entry_dict.get('wgResourceName')[0].decode('utf-8')
         
         if 'nop38a' in dn.lower():  
            #print(dn, cn, rsc, dsp)
            print(dn, dsp)
            
            if 'aunz' in dsp.lower():    
               print()
               print("FOUND AUNZ:", dn, dsp)
               print()

               '''
               dsp_ldif = modlist.modifyModlist({'wgDataStorePtr': dsp.encode('utf-8')}, {'wgDataStorePtr': [new_dsp.encode('utf-8')]})

               if(commit):
                  ldap_py.modify_s(dn, dsp_ldif)
               
               count = count + 1

               res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(aws_rsc).encode('utf-8')]})

               if(commit):
                  ldap_py.modify_s(dn, res_ldif)
               '''
            else: 
            
               '''   
               
                  print(dn, cn, rsc, dsp)

                  commit = True
                  aws_rsc = rsc.replace('norm.shared', 'norm.kcciterss')
                  new_dsp = dsp.replace('norm.shared', 'norm.kcciterss')

                  print('\t', aws_rsc)
                  print('\t', new_dsp)

                  dsp_ldif = modlist.modifyModlist({'wgDataStorePtr': dsp.encode('utf-8')}, {'wgDataStorePtr': [new_dsp.encode('utf-8')]})

                  if(commit):
                     ldap_py.modify_s(dn, dsp_ldif)

                  res_ldif = modlist.modifyModlist({'wgResourceName': rsc.encode('utf-8')}, {'wgResourceName': [(aws_rsc).encode('utf-8')]})
            
                  if(commit):
                     ldap_py.modify_s(dn, res_ldif)
               '''

      print()

   except(ldap.LDAPError) as e:
      print(e)
   finally:
      print('EXIT')


'''
##
# ldaputil MAINLINE
##
''' 
if __name__ == "__main__":
   try:

      log = open('prod_dsp_upd_0317.log', 'w')
      #log = open('qc_dsp_upd.1115.txt', 'w')

      print()
      print("ldaputil: initialize connection")
      #ldap_py = ldap.initialize('ldap://localhost:389')
      ldap_py = ldap.initialize('ldap://localhost:389')
      #ldap_py = ldap.initialize('ldap://a207947-novus-prod-use1ldap.1667.aws-int.thomsonreuters.com:636')
      ldap_py.simple_bind_s("cn=novusprod,cn=users,ou=novusaws,o=westgroup.com", "SrF7330ngErf$tdm3rdeel9")
      #ldap_py.simple_bind_s("cn=novusprod,ou=novusaws,o=westgroup.com", "SrF7330ngErf$tdm3rdeel9")
      
   #   ldap_py.simple_bind_s("cn=root,cn=qc,ou=NOVUSAWS,o=WESTGROUP.COM", "$profesSionajj47")
   #   ldap_py.simple_bind_s("cn=root,cn=qc,ou=NOVUSAWS,o=WESTGROUP.COM", "$comPP3tellymanGGdoF")
   #   ldap_py.simple_bind_s("cn=root,cn=dev,ou=novusaws,o=westgroup.com", "&comPP3tellymanGGdoF")
   #   ldap_py.simple_bind_s("cn=novusqc,cn=users,ou=novusaws,o=westgroup.com", "$profesSionajj47")
      
      #ldap_py.simple_bind_s("cn=root,cn=client,ou=novusaws,o=westgroup.com", "Ccha43mPpa.daigDINa85")
      #ldap_py.simple_bind_s("cn=root,cn=prod,ou=novusaws,o=westgroup.com", "SrF7330ngErf$tdm3rdeel9")

      #ldap_py.simple_bind_s("cn=novusclient,cn=users,ou=novusaws,o=westgroup.com", "r4raEasdoutTn0")
      #ldap_py.simple_bind_s("cn=novusprod,cn=users,ou=novusaws,o=westgroup.com", "SrF7330ngErf$tdm3rdeel9")

      print("ldaputil: ldap bound")

      buckets = 0
      file_name = "ldap_fixes_0419.xlsx"
      sheet_name = 'PROD Buckets'
      TARGET_DBS = None

      commit = True

      #update_datastore_stub(ldap_py, log, commit)
      #datastore_ptr_audit(ldap_py, log, commit)
      #update_datastore_ptrs(ldap_py, log, commit)
      #update_datastore_ptr_stub(ldap_py, log, commit)

      search_datastore_ptrs(ldap_py)

      #add_resource_entry(ldap_py, 'a207947-novus-ecpqueue-01a-db-update-prod-us-east-1', commit)
   

      #TARGET_DBS = get_xl_db_list(file_name, sheet_name)

      #update_datastores(ldap_py, log, True)
      #audit_datastore_ptrs(ldap_py, log)
      #update_datastore_ptrs(ldap_py, log, PROD, True)


   finally:

      if ldap_py is not None:
         ldap_py.unbind()
         print("ldap_dev: ldap unbound")

      if log is not None:
         log.close()

      if CONN is not None:
         CONN.close()

   '''
   resource_modify_list = {"objectclass": [b"top",b"wgResource"],
                           "cn": ["dlc".encode()],
                           "wgCurrentBuckets": ["32".encode()],
                           "wgNewBuckets": ["32".encode()]
                           }
                           
   resource_dn = 'cn=dlc,cn=Resources,cn=prod,ou=novusaws,o=westgroup.com'
   resource_ldif = modlist.addModlist(resource_modify_list)
   print('Adding: ', resource_dn)
   print('        ', resource_ldif)

   try:
      ldap_py.add_s(resource_dn, resource_ldif)
   except ldap.ALREADY_EXISTS:
      print('Entry already Exists on ldap server.', resource_dn)
   except ldap.LDAPError:
      print("Failed to add entry to the ldap server", resource_dn)
   '''
