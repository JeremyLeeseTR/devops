__author__ = "Jeremy Leese | jeremy.leese@thomsonreuters.com | Technology - Product Eng - Content Platforms"
__copyright__ = "Copyright© (C) 2022 Thomson Reuters. All Rights Reserved."
__version__ = "1.0"

import base64
from encodings import utf_8
import json
import os
import re
from configparser import ConfigParser
from mmap import ACCESS_DEFAULT
import util
import datetime

import boto3
import openpyxl
import psycopg
import psycopg2
import psycopg2.extras
import cx_Oracle
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#import util
import migmaps

LOCAL_PORT = '8899'
OPEN_SSH_TUNNEL = "cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I -c "
CLOSE_SSH_TUNNEL ="cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I --close -c "

CLIENT = 'qa-us-east-1-pg-sm'
PROD = 'prod-us-east-1-pg-sm'

CREATE_NOVUSU1_SQL = "DO $$ "
CREATE_NOVUSU1_SQL += "BEGIN "
CREATE_NOVUSU1_SQL += "CREATE ROLE novusu1 LOGIN password \'novusu1\'; "
CREATE_NOVUSU1_SQL += "EXCEPTION WHEN duplicate_object THEN RAISE NOTICE \'%, skipping\', SQLERRM USING ERRCODE = SQLSTATE; "
CREATE_NOVUSU1_SQL += "END "
CREATE_NOVUSU1_SQL += "$$;"

CREATE_NOVUSU2_SQL = "DO $$ "
CREATE_NOVUSU2_SQL += "BEGIN "
CREATE_NOVUSU2_SQL += "CREATE ROLE novusu2 LOGIN password \'novusu2\'; "
CREATE_NOVUSU2_SQL += "EXCEPTION WHEN duplicate_object THEN RAISE NOTICE \'%, skipping\', SQLERRM USING ERRCODE = SQLSTATE; "
CREATE_NOVUSU2_SQL += "END "
CREATE_NOVUSU2_SQL += "$$;"

CREATE_NOVUSR1_SQL = "DO $$ "
CREATE_NOVUSR1_SQL += "BEGIN "
CREATE_NOVUSR1_SQL += "CREATE ROLE novusr1 LOGIN password \'novusr1\'; "
CREATE_NOVUSR1_SQL += "EXCEPTION WHEN duplicate_object THEN RAISE NOTICE \'%, skipping\', SQLERRM USING ERRCODE = SQLSTATE; "
CREATE_NOVUSR1_SQL += "END "
CREATE_NOVUSR1_SQL += "$$;"

CREATE_NOVUSR2_SQL = "DO $$ "
CREATE_NOVUSR2_SQL += "BEGIN "
CREATE_NOVUSR2_SQL += "CREATE ROLE novusr2 LOGIN password \'novusr2\'; "
CREATE_NOVUSR2_SQL += "EXCEPTION WHEN duplicate_object THEN RAISE NOTICE \'%, skipping\', SQLERRM USING ERRCODE = SQLSTATE; "
CREATE_NOVUSR2_SQL += "END "
CREATE_NOVUSR2_SQL += "$$;"

NOVUSU1_LOGIN_SQL = "alter role novusu1 login;"
NOVUSU2_LOGIN_SQL = "alter role novusu2 login;"
NOVUSR1_LOGIN_SQL = "alter role novusr1 login;" 
NOVUSR2_LOGIN_SQL = "alter role novusr2 login;" 

NOVUSU1_PASSWORD_SQL = "alter role novusu1 with password \'novusu1\';"
NOVUSU2_PASSWORD_SQL = "alter role novusu2 with password \'novusu2\';"
NOVUSR1_PASSWORD_SQL = "alter role novusr1 with password \'novusr1\';"
NOVUSR2_PASSWORD_SQL = "alter role novusr2 with password \'novusr2\';" 

GRANT_RR1_SQL = "DO $do$ " 
GRANT_RR1_SQL += "DECLARE sch text; "
GRANT_RR1_SQL += "BEGIN "
GRANT_RR1_SQL += "FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE \'pg_%\' LOOP "
GRANT_RR1_SQL += "EXECUTE format($$ GRANT USAGE ON SCHEMA %I TO novusr1 $$, sch); "
GRANT_RR1_SQL += "EXECUTE format($$ GRANT SELECT ON ALL TABLES IN SCHEMA %I TO novusr1 $$, sch); "
GRANT_RR1_SQL += "EXECUTE format($$ ALTER DEFAULT PRIVILEGES IN SCHEMA %I GRANT SELECT ON TABLES TO novusr1 $$, sch); "
GRANT_RR1_SQL += "END LOOP; "
GRANT_RR1_SQL += "END; "
GRANT_RR1_SQL += "$do$;"

GRANT_RR2_SQL = "DO $do$ " 
GRANT_RR2_SQL += "DECLARE sch text; "
GRANT_RR2_SQL += "BEGIN "
GRANT_RR2_SQL += "FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE \'pg_%\' LOOP "
GRANT_RR2_SQL += "EXECUTE format($$ GRANT USAGE ON SCHEMA %I TO novusr2 $$, sch); "
GRANT_RR2_SQL += "EXECUTE format($$ GRANT SELECT ON ALL TABLES IN SCHEMA %I TO novusr2 $$, sch); "
GRANT_RR2_SQL += "EXECUTE format($$ ALTER DEFAULT PRIVILEGES IN SCHEMA %I GRANT SELECT ON TABLES TO novusr2 $$, sch); "
GRANT_RR2_SQL += "END LOOP; "
GRANT_RR2_SQL += "END; "
GRANT_RR2_SQL += "$do$;"

GRANT_RU1_SQL = "DO $do$ " 
GRANT_RU1_SQL += "DECLARE sch text; "
GRANT_RU1_SQL += "BEGIN "
GRANT_RU1_SQL += "FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE \'pg_%\' LOOP "
GRANT_RU1_SQL += "EXECUTE format($$ GRANT USAGE ON SCHEMA %I TO novusu1 $$, sch); "
GRANT_RU1_SQL += "EXECUTE format($$ GRANT SELECT ON ALL TABLES IN SCHEMA %I TO novusu1 $$, sch); "
GRANT_RU1_SQL += "EXECUTE format($$ ALTER DEFAULT PRIVILEGES IN SCHEMA %I GRANT SELECT ON TABLES TO novusu1 $$, sch); "
GRANT_RU1_SQL += "END LOOP; "
GRANT_RU1_SQL += "END; "
GRANT_RU1_SQL += "$do$;"

GRANT_RU2_SQL = "DO $do$ " 
GRANT_RU2_SQL += "DECLARE sch text; "
GRANT_RU2_SQL += "BEGIN "
GRANT_RU2_SQL += "FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE \'pg_%\' LOOP "
GRANT_RU2_SQL += "EXECUTE format($$ GRANT USAGE ON SCHEMA %I TO novusu2 $$, sch); "
GRANT_RU2_SQL += "EXECUTE format($$ GRANT SELECT ON ALL TABLES IN SCHEMA %I TO novusu2 $$, sch); "
GRANT_RU2_SQL += "EXECUTE format($$ ALTER DEFAULT PRIVILEGES IN SCHEMA %I GRANT SELECT ON TABLES TO novusu2 $$, sch); "
GRANT_RU2_SQL += "END LOOP; "
GRANT_RU2_SQL += "END; "
GRANT_RU2_SQL += "$do$;"

GRANT_UU2_SQL = "DO $do$ "
GRANT_UU2_SQL += "DECLARE "
GRANT_UU2_SQL += "sch text; " 
GRANT_UU2_SQL += "BEGIN "
GRANT_UU2_SQL += "FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE \'pg_%\' "
GRANT_UU2_SQL += "LOOP "
GRANT_UU2_SQL += "EXECUTE format($$ GRANT SELECT, INSERT, UPDATE, DELETE ON ALL TABLES IN SCHEMA %I TO novusu2 $$, sch); "
GRANT_UU2_SQL += "END LOOP; "
GRANT_UU2_SQL += "END; "
GRANT_UU2_SQL += "$do$;"

GRANT_UU1_SQL = "DO $do$ "
GRANT_UU1_SQL += "DECLARE "
GRANT_UU1_SQL += "sch text; " 
GRANT_UU1_SQL += "BEGIN "
GRANT_UU1_SQL += "FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE \'pg_%\' "
GRANT_UU1_SQL += "LOOP "
GRANT_UU1_SQL += "EXECUTE format($$ GRANT SELECT, INSERT, UPDATE, DELETE ON ALL TABLES IN SCHEMA %I TO novusu1 $$, sch); "
GRANT_UU1_SQL += "END LOOP; "
GRANT_UU1_SQL += "END; "
GRANT_UU1_SQL += "$do$;"

SKIP_SCHEMAS = [
                'public',
                'information_schema', 
                'information_schema_catalog_name',
                'applicable_roles',
                'administrable_role_authorizations',
                'collation_character_set_applicability',
                'attributes',
                'character_sets',
                'check_constraint_routine_usage',
                'column_column_usage',
                'check_constraints',
                'column_udt_usage',
                'collations',
                'column_domain_usage',
                'constraint_column_usage',
                'column_privileges',
                'columns',
                'constraint_table_usage',
                'domain_constraints',
                'domain_udt_usage',
                'domains',
                'enabled_roles',
                'key_column_usage',
                'parameters',
                'referential_constraints',
                'role_column_grants',
                'routine_privileges',
                'role_routine_grants',
                'routines',
                'schemata',
                'sequences',
                'sql_implementation_info',
                'table_constraints',
                'transforms',
                'table_privileges',
                'role_table_grants',
                'triggered_update_columns',
                'tables',
                'triggers',
                'role_udt_grants',
                'udt_privileges',
                'usage_privileges',
                'role_usage_grants',
                'user_defined_types',
                'view_column_usage',
                'view_routine_usage',
                'view_table_usage',
                'views',
                'data_type_privileges',
                '_pg_user_mappings',
                'element_types',
                '_pg_foreign_table_columns',
                'column_options',
                'user_mapping_options',
                '_pg_foreign_data_wrappers',
                'foreign_data_wrapper_options',
                'foreign_data_wrappers',
                '_pg_foreign_servers',
                'foreign_server_options',
                'user_mappings',
                'foreign_servers',
                '_pg_foreign_tables',
                'foreign_table_options',
                'foreign_tables',
                'sql_parts',
                'sql_sizing',
                'sql_features'
            ]

##
# BEGIN DEF config
# config the db for connection
##
def config(filename, section):
    # create a parser
    parser = ConfigParser()
    # read config file
    parser.read(filename)

    # get section, default to postgresql
    db = {}
    if parser.has_section(section):
        params = parser.items(section)
        for param in params:
            db[param[0]] = param[1]
    else:
        raise Exception('Section {0} not found in the {1} file'.format(section, filename))

    return db

# END DEF config

##
# BEGIN DEF test_connect
# test connect() to the db
##
def test_connect(params):
    
    conn = None

    try:

        # connect to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        conn = psycopg.connect(**params)
		
        # create a cursor
        cur = conn.cursor()
        
	    # execute a statement
        print('PostgreSQL database version:')
        cur.execute('SELECT version()')

        # display the PostgreSQL database server version
        db_version = cur.fetchone()
        print(db_version)
       
	    # close the communication with the PostgreSQL
        cur.close()

    except (Exception, psycopg.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()
            print('Database connection closed.')
        ## END if
    ## END try

## END DEF test_connect


# END DEF populate_migration_table

###
# BEGIN DEF update_cci_resources(filename)
###
def update_cci_resources(filename):
    print('BEGIN', filename)

    aws_resources = {'pgdoc05ai': 'doc.05',
                    'pgdoc06ai': 'doc.06',
                    'pgdoc02ai': 'doc.02',
                    'pgdocha01ai': 'doc.ha.01',
                    'pddoc01ai': 'doc.01'}

    try:
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8892")
        conn = psycopg.connect(dbname='pgcci01ai', user='postgres', password='9YDDqVOB[m!Mt1gV{l6{e4A77)A]U;a7', host="localhost", port="8891")
        cur = conn.cursor()

        schema_map_log = open('short_work_list.txt','r')

        schema_map = schema_map_log.readlines()
        
        for op_schema in schema_map:
            op_schema.strip()

            pos = op_schema.find('\t')

            schema = op_schema[:pos-1]
            schema.strip()
            cluster = op_schema[pos+2: len(op_schema) - 1]
            cluster.strip()

            print('s&r>' + schema + '>' + cluster + '<')
            sqls = 'SELECT collection_name, doc_resource_name, doc_schema_name FROM CCI.DOC_UPDATE_PROCESS WHERE doc_schema_name ILIKE \'' + schema + '\''
            
            sqls = 'UPDATE cci.doc_update_process SET doc_resource_name = \'' + aws_resources[cluster] + '\' WHERE doc_schema_name ILIKE \'' + schema + '\''
            cur.execute(sqls)

            sqls = 'UPDATE cci.doc SET doc_resource_name = \'' + aws_resources[cluster] + '\' WHERE doc_schema_name ILIKE \'' + schema + '\''
            cur.execute(sqls)
            
            '''
            rows = cur.fetchall()

            for row in rows:
                print(row)
            '''

            conn.commit()
            print()

    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)  
        print(sqls)

    finally:
        
        if(conn):
            conn.close()
        if(cur):
            cur.close()
        if(schema_map_log):
            schema_map_log.close()

# END DEF update_cci_resources(filename)

###
# BEGIN DEF update_cci()
###
def update_cci():
    LOG = open('qa_cci_sql.log', 'w')
    sqls = None

    try:
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8892")
        #conn = psycopg.connect(dbname='pgcci01ai', user='postgres', password='9YDDqVOB[m!Mt1gV{l6{e4A77)A]U;a7', host="localhost", port="8891")
        conn = psycopg.connect(dbname='pgcci01bq', user='postgres', password='vWbJpWKKOUGCnR{9>Hniu6}`vpH?VH?J', host="localhost", port="8898  ")
        #conn = psycopg.connect(dbname='pgcci01bq', user='postgres', password='IeIkr-KBQV}L{Ka<!}>Y,5$g4-y%Wc~u', host="localhost", port="8899")
        #cci_b_conn = psycopg.connect(dbname='pgcci01ai', user='postgres', password='9YDDqVOB[m!Mt1gV{l6{e4A77)A]U;a7', host="localhost", port="8891")
        cur = conn.cursor()

        resource_dict = {'sha.rac4.n1': 'pgdocshared01a',
                        'sha.rac4.n2': 'pgdocshared01a',
                        'SHA.RAC5.N2': 'pgdocwestlaw04a',
                        'wl.cda0320': 'pgdocwestlaw01a',
                        'wla.rac29.n1': 'pgdocwestlaw04a',
                        'wla.rac29.n2': 'pgdocwestlaw04a',
                        'WLA.RAC30.N2': 'pgdocwestlaw04a',
                        'wlnv.rac2.n1': 'pgdocwestlaw02a',
                        'wlnv.rac2.n2': 'pgdocwestlaw02a',
                        'wlnv.rac5.n1': 'pgdocwestlaw03a',
                        'WLNV.RAC5.N2': 'pgdocwestlaw03a'
                    }
        
        cluster_dict = {'pgdocshared01a': 'doc.shared.01',
                        'pgdocwestlaw04a': 'doc.westlaw.04',
                        'pgdocwestlaw01a': 'doc.westlaw.01',
                        'pgdocwestlaw02a': 'doc.westlaw.02',
                        'pgdocwestlaw03a': 'doc.westlaw.03'
                    }

        for resource in resource_dict:

            print()
            print(file=LOG)
            
            sqls = 'UPDATE cci.doc_update_process SET doc_resource_name = \'' + cluster_dict[resource_dict[resource]] + '\' WHERE doc_resource_name ILIKE \'' + resource + '\''
            print(sqls)
            print(sqls, file=LOG)
            cur.execute(sqls)
            
            sqls = 'UPDATE cci.doc SET doc_resource_name = \'' + cluster_dict[resource_dict[resource]] + '\' WHERE doc_resource_name ILIKE \'' + resource + '\''
            print(sqls)
            print(sqls, file=LOG)
            cur.execute(sqls)

            conn.commit()

    except (Exception, psycopg.DatabaseError) as error:
        print()
        print(error)  
        print(sqls)

    finally:
            
        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()

# END DEF update_cci()

###
# BEGIN get_collection_list() 
###
def get_collection_list():

    schema_map_log = open('uproc_schema_map.log','r')
    schema_map = schema_map_log.readlines()

    schema_map_dict = {}
    schema_dict_list = []

    for map in schema_map:
        #print(row, schema, cluster, file=uproc_schema_map_log)
        print(map)
        coll_name = map[2:(map.find("\', ") - 1)] 
        op_res_name = map[(map.find("\', \'") + 4):(map[map.find(", \'"):].find("\', \'") -1)]
        op_schema_name = map
        aws_schema_name = map
        aws_cluster_name = map

        schema_map_dict["coll_name"] = coll_name
        schema_map_dict["op_res_name"] = op_res_name
        schema_map_dict["op_schema_name"] = op_schema_name
        schema_map_dict["aws_schema_name"] = aws_schema_name
        schema_map_dict["aws_cluster_name"] = aws_cluster_name

        print(schema_map_dict)

        schema_dict_list = schema_map_dict
    
    return schema_dict_list

# END DEF get_collection_list() 


##
#  BEGIN DEF map_schema_locations(LOG)
##
def map_schema_locations(LOG):

    region='us-east-1'
    cur = None
    conn = None
    tunnel_open = False

    session = boto3.session.Session()
    client = session.client(service_name='secretsmanager', region_name=region)

    secrets_response = get_secrets_list(client)

    for secret in secrets_response:

        secret_name = secret['Name']

        if secret_name.find('test-us-east-1-pg-sm') >= 0:

            response = client.get_secret_value(SecretId=secret_name)

            # Decrypts secret using the associated KMS CMK.
            # Depending on whether the secret is a string or binary, one of these fields will be populated.
            if 'SecretString' in response:
                secret_s = response['SecretString']
                print('\tsecret_s:', secret_s)

            elif 'SecretBinary' in response:
                secret_b = base64.b64decode(response['SecretBinary'])
                print('\tsecret_b:', secret_b)
            else:
                print('\tSecret String NOT FOUND')

            secret_dict = json.loads(secret_s) 

            cluster_id = secret_dict['dbClusterIdentifier'] 
            host = secret_dict['host']
            db_name = secret_dict['dbname']
            user = secret_dict['username']
            password = secret_dict['password']
            port = secret_dict['port']

            # ssh tunnel to aws pg db
            tunnel_cmd = OPEN_SSH_TUNNEL + host

            print()
            print("Exec CMD: " + tunnel_cmd)
            os.system(tunnel_cmd)
            tunnel_open = True

            print()
            print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', port, '...')

            try:
                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=port)
            except(psycopg.DatabaseError) as error:
                print(error)
                continue

            try:
                cur = conn.cursor()
                cur.execute("SELECT schema_name FROM information_schema.schemata WHERE " +
                            "(schema_name NOT LIKE \'pg_%\' " +
                            "AND schema_name NOT LIKE \'repack%\' " +
                            "AND schema_name NOT LIKE \'partman%\' " +
                            "AND schema_name NOT LIKE \'public%\' " +
                            "AND schema_name NOT LIKE \'information_schema%\' " +
                            "AND schema_name NOT LIKE \'aws_%\')")

                print()
                schemas = cur.fetchall()

                for schema in schemas:
                    print(db_name + ',' + schema[0])  
                    print(db_name + ',' + schema[0], file=LOG)           

            except (Exception, psycopg.DatabaseError) as error:
                print(error)
            finally:
                if cur is not None:
                    cur.close()
                    cur = None
                    print('Database cursor closed.')
                if conn is not None:
                    conn.close()
                    conn = None
                    print('Database connection closed.')
                if tunnel_open:
                    tunnel_cmd = CLOSE_SSH_TUNNEL + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
    
    if cur is not None:
        cur.close()
        print('Database cursor closed.')
    if conn is not None:
        conn.close()
        print('Database connection closed.')
    if tunnel_open:
        tunnel_cmd = CLOSE_SSH_TUNNEL + host
        os.system(tunnel_cmd)
        tunnel_open = False

#  BEGIN DEF map_schema_locations(LOG)

###
# BEGIN DEF map_schema_to_op_rsc(log)
###
def map_schema_to_op_rsc(log):
    print("BEGIN: map_schema_to_op_rsc(log)")

    filename = "test_py_xl.xlsx"
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row
    aws_schema = None

    db_name = 'pgcci01bi'
    db_user = 'postgres' 
    db_password = 'm8ZV*>IK{JOlG_e&4aAD8_uD?pOVaCtL' 

    try:
        conn = psycopg.connect(dbname=db_name, user=db_user, password=db_password, host="localhost", port="8898")
        cur = conn.cursor()
    
        for i in range(1, max_row):
            aws_schema = ws.cell(row = i, column = 2).value
            aws_db = ws.cell(row = i, column = 1).value

            #print('aws_schema:', type(aws_schema), aws_schema)
            #print('aws_db:', type(aws_db), aws_db)
            #print('aws_schema:', aws_schema)

            sqls = 'SELECT doc_resource_name FROM cci.doc_update_process WHERE doc_schema_name ILIKE \'' + aws_schema + '\''
            #print(sqls)
            #print(sqls, file=log)
            cur.execute(sqls)

            rsc_names = cur.fetchall()
            #print('rsc_names:', type(rsc_names), rsc_names)
            for name in rsc_names:
                #print(aws_db, aws_schema, name[0])  
                print(aws_db, aws_schema, name[0], file=log)  

        conn.commit()
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close

# END DEF map_schema_to_op_rsc(log)

###
# BEGIN DEF uproc_schema_map(log)
# returns a list of dictionaries containing the OP to AWS schema mmapping
###
def uproc_schema_map(log):
    print('BEGIN', log)

    aws_resources = {'pgdoc05ai': 'doc.05',
                    'pgdoc06ai': 'doc.06',
                    'pgdoc02ai': 'doc.02',
                    'pgdocha01ai': 'doc.ha.01',
                    'pddoc01ai': 'doc.01'}

    try:
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8892")
        conn = psycopg.connect(dbname='pgcci01ai', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8891")
        conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8894")
        #conn = psycopg.connect(dbname='pgcci01aq', user='postgres', password='IeIkr-KBQV}L{Ka<!}>Y,5$g4-y%Wc~u', host="localhost", port="8899")
        #cci_b_conn = psycopg.connect(dbname='pgcci01ai', user='postgres', password='9YDDqVOB[m!Mt1gV{l6{e4A77)A]U;a7', host="localhost", port="8891")
        cur = conn.cursor()

        schema_map_log = open('aws_qc_schemas.txt','r')
        upd_dat_log = open('aws_qc_upd_dat.txt','r')
        uproc_schema_map_log = open('uproc_schema_map.log.3','w')

        schema_map = schema_map_log.readlines()
        upd_dat_map = upd_dat_log.readlines()
        schema_dict_list = []
        
        for op_schema in schema_map:
            op_schema.strip()

            pos = op_schema.find('\t')

            cluster = op_schema[:pos-1]
            cluster.strip()
            schema = op_schema[pos+1: len(op_schema) - 1]
            schema.strip()
    
            for upd_dat in upd_dat_map:
                upd_dat.strip()

                if re.search(schema, upd_dat, re.IGNORECASE):
                    
                    sqls = 'SELECT collection_name, doc_resource_name, doc_schema_name FROM CCI.DOC_UPDATE_PROCESS WHERE doc_schema_name ILIKE \'' + schema + '\''
                    print(sqls)

                    '''
                    sqls = 'UPDATE cci.doc_update_process SET doc_resource_name = \'' + aws_resources[cluster] + '\' WHERE doc_schema_name ILIKE \'' + schema + '\''
                    cur.execute(sqls)

                    sqls = 'UPDATE cci.doc SET doc_resource_name = \'' + aws_resources[cluster] + '\' WHERE doc_schema_name ILIKE \'' + schema + '\''
                    
                    '''

                    cur.execute(sqls)
                    rows = cur.fetchall()
                    schema_map_dict = {}
                    
                    for row in rows:
                        
                        schema_map_dict["coll_name"] = row[0]
                        schema_map_dict["op_res_name"] = row[1]
                        schema_map_dict["op_schema_name"] = row[2]
                        schema_map_dict["aws_schema_name"] = schema
                        schema_map_dict["aws_cluster_name"] = cluster

                        print(schema_map_dict, file=uproc_schema_map_log)

                        schema_dict_list.append(schema_map_dict)

        return schema_dict_list

    except (Exception, psycopg.DatabaseError) as error:
                # skip duplicates
                print()
                print(error)  
                print(sqls)
    finally:
        
        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()
        if(schema_map_log is not None):
            schema_map_log.close()

# END DEF uproc_schema_map(log)

##
# BEGIN DEF grant_read_access(user_name, conn)
##
def grant_read_access(user_name, conn):

    try:
        cur = conn.cursor()

        print("Granting read access...")   
        sqls = "DO $do$ "
        sqls += "DECLARE "
        sqls += "sch text; "
        sqls += "BEGIN "
        sqls += "    FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE 'pg_%'"
        sqls += "    LOOP "
        sqls += "        EXECUTE format($$ GRANT USAGE ON SCHEMA %I TO " + user_name + " $$, sch); "
        sqls += "        EXECUTE format($$ GRANT SELECT ON ALL TABLES IN SCHEMA %I TO  " + user_name + " $$, sch); "
        sqls += "        EXECUTE format($$ ALTER DEFAULT PRIVILEGES IN SCHEMA %I GRANT SELECT ON TABLES TO  " + user_name + " $$, sch); "
        sqls += "    END LOOP; "
        sqls += "END; "
        sqls += "$do$;"

        print()
        print(sqls)
        print()
        cur.execute(sqls)
    
        cur.execute("GRANT pg_read_all_settings TO " + user_name)

        # save commit until the end. getting stopped mid process can lead to bad things
        conn.commit()

        cur.close()
    except (Exception, psycopg.DatabaseError) as error:
        print("DB ERROR:", error)
        raise Exception


# END DEF grantReadAccess()

##
#   BEGION DEF grant_datadog_access
##
def grant_datadog_access():

    cur = None
    conn = None
    tunnel_open = False

    #db_instances = util.get_db_instances('us-east-1')
    
    db_instances = []
 
    for db_dict in db_instances:
        endpoint = (db_dict.get('Endpoint'))
        address = (endpoint.get('Address')) 

        # very kludgey. use clusters, and a better way to id dbs targeted for change
        if '-dev-pg-cluster-' in address:

            db_user = 'datadog'
            db_password = 'datadog'
            
            
            db_name = db_dict.get('DBName')
            cluster_host = address.replace('us-east-1-inst', 'us-east-1')

            db_name = 'pgdoc01ad'

            # ssh tunnel to aws pg db
            tunnel_cmd = OPEN_SSH_TUNNEL + cluster_host

            tunnel_cmd = OPEN_SSH_TUNNEL + 'a207947-doc-01a-dev-pg-cluster-us-east-1.cluster-coypymqvk3us.us-east-1.rds.amazonaws.com'

            print()
            print("Exec CMD: " + tunnel_cmd)
            os.system(tunnel_cmd)
            tunnel_open = True

            print()
            print('Connecting to the PostgreSQL database ' + db_name + '...')
            conn = psycopg.connect(dbname=db_name, user=db_user, password=db_password, host="localhost", port="5432")

            try:
                cur = conn.cursor()
                cur.execute("GRANT pg_read_all_settings TO datadog")
                cur.execute("GRANT SELECT ON pg_stat_replication TO datadog")

            except (Exception, psycopg.DatabaseError) as error:
                print(error)
            finally:
                if cur is not None:
                    cur.close()
                    cur = None
                    print('Database cursor closed.')
                if conn is not None:
                    conn.close()
                    conn = None
                    print('Database connection closed.')
                if tunnel_open:
                    tunnel_cmd = CLOSE_SSH_TUNNEL + cluster_host
                    print("Exec CMD: " + tunnel_cmd)
                    os.system(tunnel_cmd)
                    tunnel_open = False

    if cur is not None:
        cur.close()
        print('Database cursor closed.')
    if conn is not None:
        conn.close()
        print('Database cursor closed.')
    if tunnel_open:
        tunnel_cmd = CLOSE_SSH_TUNNEL + cluster_host
        print("Exec CMD: " + tunnel_cmd)
        os.system(tunnel_cmd)
        tunnel_open = False

#   END DEF grant_datadog_access

##
#  BEGIN DEF grant_update_access(user_name, conn)
##
def grant_update_access(user_name, conn):

    cur = conn.cursor()

    # grant read    
    grant_read_access(user_name, conn)

    print("Granting update access...")

    try:
        # grant update/write
        cur.execute("ALTER USER " + user_name + " CREATEDB")

        sqls = "DO $do$ "
        sqls += "DECLARE "
        sqls += "sch text; "
        sqls += "BEGIN "
        sqls += "    FOR sch IN SELECT nspname FROM pg_namespace where nspname NOT LIKE 'pg_%'"
        sqls += "    LOOP "
        sqls += "        EXECUTE format($$ GRANT SELECT, INSERT, UPDATE, DELETE ON ALL TABLES IN SCHEMA %I TO " + user_name + " $$, sch); "
        sqls += "    END LOOP; "
        sqls += "END; "
        sqls += "$do$;"

        print()
        print(sqls)
        print()

        cur.execute(sqls)
        conn.commit()
        cur.close()

    except (Exception, psycopg.DatabaseError) as error:
        
        print("DB ERROR:", error)
        raise Exception


#  END DEF grant_update_access(user_name, conn)

##
#  BEGIN DEF fix_function(commit)
##
def fix_function(commit):
    try:
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8892")
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8891")
        #conn = psycopg.connect(dbname='pgcci01aq', user='postgres', password='IeIkr-KBQV}L{Ka<!}>Y,5$g4-y%Wc~u', host="localhost", port="8899")
        #conn = psycopg.connect(dbname='pgcci01aq', user='novusu2', password='novusu2', host="localhost", port="8891")
        #conn = psycopg.connect(dbname='pgcci01bq', user='novusu2', password='novusu2', host="localhost", port="8894")
        #conn = psycopg.connect(dbname='pgnimsccscaleretr01ai', user='postgres', password='}yEnG#$}ZeJwo>[xg)59,9?HW{pqF`1v', host="localhost", port="8891")
        
        log_file = None
        conn = None
        cur = None

        region='us-east-1'
        cur = None
        conn = None
        tunnel_open = False

        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)

        secrets_response = get_secrets_list(client)

        for secret in secrets_response:

            secret_name = secret['Name']

            if secret_name.find('qa-us-east-1-pg-sm') >= 0 and secret_name.find('nims') >=0:

                response = client.get_secret_value(SecretId=secret_name)

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']
                    print('\tsecret_s:', secret_s)

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                    print('\tsecret_b:', secret_b)
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                #cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                #port = secret_dict['port']

                # ssh tunnel to aws pg db
                tunnel_cmd = OPEN_SSH_TUNNEL + host

                print()
                print("Exec CMD: " + tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True

                print()
                print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
                print()

                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                #grant_update_access('novusu1', conn)
                #grant_read_access('novusr1', conn)
                #log_file = open('client_schema_audit.1209.txt', 'r')
                
                schema_name = None
                

                sqls = "SELECT nspname FROM pg_namespace where nspname NOT LIKE 'pg_%'"

                print()
                print(sqls)
                
                cur = conn.cursor()
                result = cur.execute(sqls)
                conn.commit()

                rows = cur.fetchall()

                for row in rows:

                    schema_name = row[0]

                    sqls =  "CREATE OR REPLACE FUNCTION " + schema_name
                    sqls += ".\"sidetab_del_t$cms_content\"() "
                    sqls +=     "RETURNS trigger "
                    sqls +=     "LANGUAGE plpgsql "
                    sqls +=     "AS $function$ "
                    
                    sqls +=     "BEGIN "
                    sqls +=         "DELETE FROM " + schema_name + ".cms_multiple "
                    sqls +=             "WHERE old.root_id = " + schema_name + ".cms_multiple.root_id; "
                    sqls +=                 "RETURN OLD; "
                    sqls +=     "END; "
                    sqls += "$function$;"

                    print('SQL:', sqls)
                    cur = conn.cursor()
                    result = cur.execute(sqls)

                    if commit:
                        conn.commit()

                    print('schema_name: ' + schema_name)
                
                print()
                print('db_name:',db_name)
                print()

                if(log_file is not None):
                    log_file.close()
                if(conn is not None):
                    conn.close()
                if(cur is not None):
                    cur.close()
                if tunnel_open:
                    tunnel_cmd = CLOSE_SSH_TUNNEL + host
                    os.system(tunnel_cmd)
                    tunnel_open = False

    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)  
    finally:
        
        if(log_file is not None):
            log_file.close()
        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()

#  END DEF fix_function(commit)

##
#  BEGIN DEF get_add_db_user_sql(user, password)
##
def get_add_db_user_sql(user, password):

    sql =  "DO "
    sql += "$do$ "
    sql += "    BEGIN "
    sql += "    IF NOT EXISTS ("
    sql += "        SELECT FROM pg_catalog.pg_roles  "
    sql += "        WHERE  rolname = \'" + user + "\') THEN "
    sql += "        CREATE ROLE " + user + " LOGIN PASSWORD \'" + password + "\';"
    sql += "    END IF;"
    sql += "END"
    sql += "$do$;"

    return sql

#  END DEF get_add_db_user_sql(user, password)

# END DEF grant_access_oneoff()

##
#  BEGIN DEF get_secrets_list()
##
def get_secrets_list(client):

    secrets_list = None
    next_token = None

    try:
        response = client.list_secrets(MaxResults=100)
        secrets_list = response.get('SecretList')
        next_token = response.get('NextToken')

        while(next_token):
            response = client.list_secrets(MaxResults=100, NextToken=next_token)
            secrets_list.extend(response.get('SecretList'))
            next_token = response.get('NextToken')

        return secrets_list
    
    except ClientError as e:
        raise Exception("boto3 client error in get_secrets: " + e.__str__())
    except Exception as e:
        raise Exception("Unexpected error in get_secrets: " + e.__str__()) 

# END DEF get_secrets_list()

##
#  BEGIN DEF get_cci_schema_data()
##
def get_cci_schema_data():
    conn = None
    cur = None

    conn = psycopg.connect(dbname='pgcci01aq', user='postgres', password='1H.9,<oN9mOfq<2|TZua62,W]N(Yc7_C', host="localhost", port='1234')
    cur = conn.cursor()
    
    schema_map_file = open('schema_map_toc.txt', 'w')

    schema_name = None
    resource_name = None
    collection_name = None
    
    with open('client_schemas.txt', 'r') as schema_file:
        
        for schema_name in schema_file:
            schema_name = schema_name.rstrip()
            
            #sqls = "SELECT doc_resource_name, doc_schema_name FROM cci.doc_update_process WHERE doc_schema_name LIKE \'" + schema_name + "\'"
            #sqls = "SELECT meta_resource_name, meta_schema_name FROM cci.mm_update_proc WHERE meta_schema_name = \'" + schema_name + "\'"
            sqls = "SELECT toc_resource_name, toc_schema_name FROM cci.toc_update_process WHERE toc_schema_name = \'" + schema_name + "\'"
            
            cur.execute(sqls)
        
            rows = cur.fetchall()

            for row in rows:
                
                resource_name = row[0]
                schema_name = row[1]

                print(resource_name, schema_name)
                print(resource_name, schema_name, file=schema_map_file)
    
    if cur is not None:
        cur.close()
    if conn is not None:
        conn.close()
    if schema_map_file is not None:
        schema_map_file.close()

# END DEF get_secrets_list()

##
# BEGIN DEF fix_trigger_functions(commit)
##
def fix_trigger_functions(commit):
    try:

        log_file = None
        conn = None
        cur = None

        region='us-east-1'
        cur = None
        conn = None
        tunnel_open = False

        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)

        secrets_response = get_secrets_list(client)

        for secret in secrets_response:

            secret_name = secret['Name']

            #print('secret_name:', secret_name)

            if secret_name.find('test-us-east-1-pg-sm') >= 0 and secret_name.find('nims') >=0:

                response = client.get_secret_value(SecretId=secret_name)

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']
                    print('\tsecret_s:', secret_s)

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                    print('\tsecret_b:', secret_b)
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                #cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                #port = secret_dict['port']

                # ssh tunnel to aws pg db
                tunnel_cmd = OPEN_SSH_TUNNEL + host

                print()
                print("Exec CMD: " + tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True

                print()
                print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
                print()

                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                sql = "SELECT * FROM information_schema.routines "
                sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t$%nims_payload\';"
                print(sql)

                cur = conn.cursor()
                cur.execute(sql)
                conn.commit()
                schema_routines = cur.fetchall()

                print('db_name:',db_name)
                
                for schema_routine in schema_routines:
                    routine_catalog = schema_routine[3]
                    routine_schema = schema_routine[4]
                    routine_name = schema_routine[5]
                    routine_definition = schema_routine[37]

                    #print (routine_definition)
                    #new_routine_definition = routine_definition.replace('oldrow.root_id', 'old.root_id')
                    #print(new_routine_definition) 

                    '''
                    sql = "CREATE OR REPLACE FUNCTION " + routine_schema + ".\"sidetab_del_t$nims_payload\"() "
                    sql += " RETURNS trigger "
                    sql += " LANGUAGE plpgsql "
                    sql += "AS $function$ "
                    sql += "BEGIN "
                    sql += "DELETE FROM " + routine_schema + ".flat_side_tab "
                    sql += "WHERE old.root_id = " + routine_schema + ".flat_side_tab.root_id; "
                    sql += "DELETE FROM " + routine_schema + ".multiple "
                    sql += "WHERE old.root_id = " + routine_schema + ".multiple.root_id; "
                    sql += "DELETE FROM " + routine_schema + ".n_view "
                    sql += "WHERE old.root_id = " + routine_schema + ".n_view.root_id; "
                    sql += "RETURN OLD; "
                    sql += "END; "
                    sql += "$function$;"
                    '''

                    sql = "CREATE OR REPLACE FUNCTION " + routine_schema + ".\"sidetab_del_t$nims_payload\"() "
                    sql += " RETURNS trigger "
                    sql += " LANGUAGE plpgsql "
                    sql += "AS $function$ "
                    sql += "BEGIN "
                    sql += "DELETE FROM " + routine_schema + ".flat_side_tab "
                    sql += "WHERE old.root_id = " + routine_schema + ".flat_side_tab.root_id; "
                    sql += "RETURN OLD; "
                    sql += "END; "
                    sql += "$function$;"
                    #sql += "ALTER FUNCTION " + routine_schema + ".\"sidetab_del_t$relationship\"() OWNER TO postgres;"
                    #sql += "GRANT ALL ON FUNCTION " + routine_schema + ".\"sidetab_del_t$relationship\"() TO postgres;"

                    print(sql)

                    cur = conn.cursor()
                    cur.execute(sql)
                        
                    if(commit):
                        conn.commit()

                    

                    '''
                    if routine_name.find('cms_content') >= 0:
                        print('\t', routine_schema, ':', routine_name)
                        sql = "DROP FUNCTION IF EXISTS " + routine_schema + ".sidetab_del_t$cms_content;"
                        print('\tsql:', sql)

                        cur = conn.cursor()
                        cur.execute(sql)
                        
                        if(commit):
                            conn.commit()
                    '''

                    #print('\troutine_definition:', routine_definition)

                if(log_file is not None):
                    log_file.close()
                if(conn is not None):
                    conn.close()
                if(cur is not None):
                    cur.close()
                if tunnel_open:
                    tunnel_cmd = CLOSE_SSH_TUNNEL + host
                    os.system(tunnel_cmd)
                    tunnel_open = False

    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)  
    finally:
        
        if(log_file is not None):
            log_file.close()
        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()

# END DEF fix_trigger_functions(commit)

##
#  BEGIN DEF grant_user_access(service)
##
def grant_user_access(service, start_secret):
    try:
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8892")
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8891")
        #conn = psycopg.connect(dbname='pgcci01aq', user='postgres', password='IeIkr-KBQV}L{Ka<!}>Y,5$g4-y%Wc~u', host="localhost", port="8899")
        #conn = psycopg.connect(dbname='pgcci01aq', user='novusu2', password='novusu2', host="localhost", port="8891")
        #conn = psycopg.connect(dbname='pgcci01bq', user='novusu2', password='novusu2', host="localhost", port="8894")
        #conn = psycopg.connect(dbname='pgnimsccscaleretr01ai', user='postgres', password='}yEnG#$}ZeJwo>[xg)59,9?HW{pqF`1v', host="localhost", port="8891")

        conn = None
        cur = None

        region='us-east-1'
        cur = None
        conn = None
        tunnel_open = False

        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)
        
        secrets_response = get_secrets_list(client)

        skip = True

        for secret in secrets_response:

            secret_name = secret['Name']

            if 'qa-us-east-1-pg-sm' in secret_name and \
                    service in secret_name and \
                    'z-prod' not in secret_name and \
                    'docloc' not in secret_name and \
                    'dlc' not in secret_name:

                if (start_secret in secret_name or not skip):
                    skip = False

                    response = client.get_secret_value(SecretId=secret_name)

                    # Decrypts secret using the associated KMS CMK.
                    # Depending on whether the secret is a string or binary, one of these fields will be populated.
                    if 'SecretString' in response:
                        secret_s = response['SecretString']
                        print('\tsecret_s:', secret_s)

                    elif 'SecretBinary' in response:
                        secret_b = base64.b64decode(response['SecretBinary'])
                        print('\tsecret_b:', secret_b)
                    else:
                        print('\tSecret String NOT FOUND')

                    secret_dict = json.loads(secret_s) 

                    #cluster_id = secret_dict['dbClusterIdentifier'] 
                    host = secret_dict['host']
                    db_name = secret_dict['dbname']
                    user = secret_dict['username']
                    password = secret_dict['password']
                    #port = secret_dict['port']

                    # ssh tunnel to aws pg db
                    tunnel_cmd = OPEN_SSH_TUNNEL + host

                    reconnect = False

                    try:
                        print()
                        print("Exec CMD: " + tunnel_cmd)
                        os.system(tunnel_cmd)
                        tunnel_open = True

                        print()
                        print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
                        print()

                        conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                        try:
                            sql = get_add_db_user_sql('novusu1', 'novusu1')

                            print('SQL:', sql)
                            cur = conn.cursor()
                            result = cur.execute(sql)
                            conn.commit()
                            
                            grant_update_access('novusu1', conn)

                        except (Exception, psycopg.DatabaseError) as error:
                            # skip not founds
                            print()
                            print(error) 
                            
                            if(cur is not None):
                                cur.close()
                            if(conn is not None):
                                conn.close()
                            
                            conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)
                        
                        try:
                            sql = get_add_db_user_sql('novusu2', 'novusu2')

                            print('SQL:', sql)
                            cur = conn.cursor()
                            result = cur.execute(sql)
                            conn.commit()

                            grant_update_access('novusu2', conn)
                            #log_file = open('client_schema_audit.1209.txt', 'r')
                        except (Exception, psycopg.DatabaseError) as error:
                            # skip not founds
                            print()
                            print(error)
                            print(error, file=error_log) 
                            
                            if(cur is not None):
                                cur.close()
                            if(conn is not None):
                                conn.close()
                            
                            conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                        try:
                            sql = get_add_db_user_sql('novusr1', 'novusr1')

                            print('SQL:', sql)
                            cur = conn.cursor()
                            result = cur.execute(sql)
                            conn.commit()

                            grant_read_access('novusr1', conn)
                            #log_file = open('client_schema_audit.1209.txt', 'r')
                        except (Exception, psycopg.DatabaseError) as error:
                            # skip not founds
                            print()
                            print(error)
                            print(error, file=error_log) 
                            
                            if(cur is not None):
                                cur.close()
                            if(conn is not None):
                                conn.close()
                            
                            conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)
                        
                        if(reconnect):
                            
                            reconnect = False

                        try:
                            sql = get_add_db_user_sql('novusr2', 'novusr2')

                            print('SQL:', sql)
                            cur = conn.cursor()
                            result = cur.execute(sql)
                            conn.commit()

                            grant_read_access('novusr2', conn)
                            #log_file = open('client_schema_audit.1209.txt', 'r')
                            
                        except (Exception, psycopg.DatabaseError) as error:
                            # skip not founds
                            print()
                            print(error)
                            print(error, file=error_log)

                            if(cur is not None):
                                cur.close()
                            if(conn is not None):
                                conn.close()
                            
                            conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                        if(cur is not None):
                            cur.close()
                        if(conn is not None):
                            conn.close()
                        
                        if tunnel_open:
                            tunnel_cmd = CLOSE_SSH_TUNNEL + host
                            os.system(tunnel_cmd)
                            tunnel_open = False
                    
                    except (Exception, psycopg.DatabaseError) as error:
                            # skip not founds
                            print()
                            print(error)
                            
                            if(cur is not None):
                                cur.close()
                            if(conn is not None):
                                conn.close()
                    
                    
    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)  
    finally:

        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()

#  END DEF grant_user_access(service)

##
#  BEGIN DEF fix_trigger_functions(commit)
##
def fix_trigger_functions(commit):
    try:

        log_file = None
        conn = None
        cur = None

        region='us-east-1'
        cur = None
        conn = None
        tunnel_open = False

        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)

        secrets_response = get_secrets_list(client)

        for secret in secrets_response:

            secret_name = secret['Name']

            #print('secret_name:', secret_name)

            if secret_name.find('test-us-east-1-pg-sm') >= 0 and secret_name.find('nims') >=0:

                response = client.get_secret_value(SecretId=secret_name)

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']
                    print('\tsecret_s:', secret_s)

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                    print('\tsecret_b:', secret_b)
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                #cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                #port = secret_dict['port']

                # ssh tunnel to aws pg db
                tunnel_cmd = OPEN_SSH_TUNNEL + host

                print()
                print("Exec CMD: " + tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True

                print()
                print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
                print()

                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                sql = "SELECT * FROM information_schema.routines "
                sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t$%nims_payload\';"
                print(sql)

                cur = conn.cursor()
                cur.execute(sql)
                conn.commit()
                schema_routines = cur.fetchall()

                print('db_name:',db_name)
                
                for schema_routine in schema_routines:
                    routine_catalog = schema_routine[3]
                    routine_schema = schema_routine[4]
                    routine_name = schema_routine[5]
                    routine_definition = schema_routine[37]

                    print (routine_definition)
                    #new_routine_definition = routine_definition.replace('oldrow.root_id', 'old.root_id')
                    #print(new_routine_definition) 

                    '''
                    sql = "CREATE OR REPLACE FUNCTION " + routine_schema + ".\"sidetab_del_t$nims_payload\"() "
                    sql += " RETURNS trigger "
                    sql += " LANGUAGE plpgsql "
                    sql += "AS $function$ "
                    sql += "BEGIN "
                    sql += "DELETE FROM " + routine_schema + ".flat_side_tab "
                    sql += "WHERE old.root_id = " + routine_schema + ".flat_side_tab.root_id; "
                    sql += "DELETE FROM " + routine_schema + ".multiple "
                    sql += "WHERE old.root_id = " + routine_schema + ".multiple.root_id; "
                    sql += "DELETE FROM " + routine_schema + ".n_view "
                    sql += "WHERE old.root_id = " + routine_schema + ".n_view.root_id; "
                    sql += "RETURN OLD; "
                    sql += "END; "
                    sql += "$function$;"
                    '''

                    sql = "CREATE OR REPLACE FUNCTION " + routine_schema + ".\"sidetab_del_t$nims_payload\"() "
                    sql += " RETURNS trigger "
                    sql += " LANGUAGE plpgsql "
                    sql += "AS $function$ "
                    sql += "BEGIN "
                    sql += "DELETE FROM " + routine_schema + ".flat_side_tab "
                    sql += "WHERE old.root_id = " + routine_schema + ".flat_side_tab.root_id; "
                    sql += "RETURN OLD; "
                    sql += "END; "
                    sql += "$function$;"
                    #sql += "ALTER FUNCTION " + routine_schema + ".\"sidetab_del_t$relationship\"() OWNER TO postgres;"
                    #sql += "GRANT ALL ON FUNCTION " + routine_schema + ".\"sidetab_del_t$relationship\"() TO postgres;"

                    '''
                    print(sql)
                    cur = conn.cursor()
                    cur.execute(sql)
                        
                    if(commit):
                        conn.commit()
                    '''
                    

                    '''
                    if routine_name.find('cms_content') >= 0:
                        print('\t', routine_schema, ':', routine_name)
                        sql = "DROP FUNCTION IF EXISTS " + routine_schema + ".sidetab_del_t$cms_content;"
                        print('\tsql:', sql)

                        cur = conn.cursor()
                        cur.execute(sql)
                        
                        if(commit):
                            conn.commit()
                    '''

                    #print('\troutine_definition:', routine_definition)

                if(log_file is not None):
                    log_file.close()
                if(conn is not None):
                    conn.close()
                if(cur is not None):
                    cur.close()
                if tunnel_open:
                    tunnel_cmd = CLOSE_SSH_TUNNEL + host
                    os.system(tunnel_cmd)
                    tunnel_open = False

    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)  
    finally:
        
        if(log_file is not None):
            log_file.close()
        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()

#  END DEF fix_trigger_functions(commit)

##
#  BEGIN DEF add_trigger_functions(db_pattern)
##
def add_trigger_functions(db_pattern, port, commit):

    region = 'us-east-1'
    conn = None
    cur = None
    tunnel_open = False

    skip_schemas = [
                        'datadog',
                        'heartbeat_monitor',
                        'migration_admin',
                        'perfstat',
                        'appqossys',
                        'stdbyperf',
                        'stdbyperf',
                        'perfstat',
                        'sys',
                        'wmsys',
                        'outln',
                        'tsmsys',
                        'system',
                        'migration_admin',
                        'information_schema',
                        'public',
                        'partman',
                        'repack',
                        'sys.whr',
                        'aws_oracle_context',
                        'aws_oracle_data',
                        'aws_oracle_ext',
                        'aws_postgis',
                        'aws',
                        'dbsnmp',
                        'killsession_user',
                        'dba_utls',
                        'aws_oracle_context'
                    ]

    try:
        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)

        log_file = os.getcwd() + '\\log\\trigger_audit.log'
        log = open(log_file, 'w')

        secrets_response = get_secrets_list(client)

        for secret in secrets_response:

            secret_name = secret['Name']
            print('.', end='')
            if 'qa' in secret_name and db_pattern in secret_name and 'a-qa-us' in secret_name and 'z-q' not in secret_name:

                try:
                    response = client.get_secret_value(SecretId=secret_name)
                except (Exception, psycopg.DatabaseError) as error:
                    # skip duplicates
                    continue

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                #cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                #port = secret_dict['port']

                open_tunnel = "cloud-tool ssh-tunnel -r " + str(port) + " -I -c "
                close_tunnel ="cloud-tool ssh-tunnel -r " + str(port) + " -I --close -c "

                # ssh tunnel to aws pg db
                tunnel_cmd = open_tunnel + host
                print(tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True

                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=str(port))

                schema_sql = "SELECT nspname FROM pg_namespace where nspname NOT LIKE \'pg_%\'"

                cur = conn.cursor()
                cur.execute(schema_sql)
                conn.commit()
                schemas = cur.fetchall()

                for schema in schemas:
        
                    schema_name = schema[0]

                    if schema_name in skip_schemas:
                        print('.', end='')

                    else:

                        table_sql = "SELECT table_name FROM information_schema.tables WHERE table_schema=\'" + schema_name + "\';"

                        cur.execute(table_sql)
                        conn.commit()
                        schema_tables = cur.fetchall()

                        table_name = None

                        sql = "SELECT routine_schema, routine_name, routine_definition FROM information_schema.routines "
                        sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t%\' AND routine_schema=\'" + schema_name + "\';"
                        
                        #print(sql)

                        cur = conn.cursor()
                        cur.execute(sql)
                        conn.commit()
                        schema_routines = cur.fetchall()

                        if len(schema_routines) > 0:
                            print('.', end='')
                            #print('FOUND routine_definition for:', schema_name)
                            #for schema_routine in schema_routines:
                            #    print(schema_routine[1])

                        else:            
                            # create trigger function
                            #print()
                            #print('CHECK CREATE Function For:', schema_name)

                            for table in schema_tables:
                                #print('\t' + table[0])
                                table_name = table[0]


                                column_select = "SELECT attrelid::regclass AS tbl, "
                                column_select += " attname AS col, "
                                column_select += " atttypid::regtype AS datatype "
                                column_select += "FROM  pg_attribute "
                                column_select += "WHERE attrelid = \'" + schema_name + "." + table_name + "\'::regclass "
                                column_select += "AND attnum > 0 "
                                column_select += "AND NOT attisdropped "
                                column_select += "ORDER BY attnum; \n"

                                #column_plan = "PREPARE colsplan (text, text) AS \n" 
                                #column_plan += column_select
                                #column_plan += "EXECUTE colsplan('" + schema_name + "\', \'" + table_name + "\');" 


                                cur = conn.cursor()
                                cur.execute(column_select)
                                conn.commit()
                                columns = cur.fetchall()

                                collection_id_found = False
                                for column in columns:
                                    #print('\t\t' + column[1])
                                    if 'collection_id' in column[1]:
                                        collection_id_found = True

                                if (not collection_id_found 
                                    and 'xml_equivalency_r' not in table_name 
                                    and 'xml_equivalency_l' not in table_name):

                                    print('collection_id column NOT FOUND:', db_name, schema_name, ':', table[0])

                                    routine_definition = "BEGIN\n    "
                                    routine_definition += "DELETE FROM " + schema_name + "." + table_name 
                                    routine_definition += " WHERE old.root_id = " + schema_name + "." + table_name + ".root_id;\n"
                                    routine_definition += "RETURN OLD;\n"
                                    routine_definition += 'END;\n'

                                    routine_name = '\"sidetab_del_t$nims_payload\"()'

                                    function_sql = "CREATE FUNCTION " + schema_name + "." + routine_name + "() "
                                    function_sql += " RETURNS trigger "
                                    function_sql += " LANGUAGE plpgsql "
                                    function_sql += "AS $function$\n"
                                    function_sql += routine_definition
                                    function_sql += "$function$;"

                                    print(function_sql)

                                    #print(function_sql)
                                    #print()

                                    #cur.execute(function_sql)
                                    #conn.commit()

                if(conn is not None):
                    conn.close()
                if(cur is not None):
                    cur.close()
                if tunnel_open:
                    tunnel_cmd = close_tunnel + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
    
    except (Exception, psycopg.DatabaseError) as error:

        print(error)
        print(error.with_traceback)

    finally:

        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()
        if tunnel_open:
            tunnel_cmd = close_tunnel + host
            os.system(tunnel_cmd)
            tunnel_open = False

#  END DEF add_trigger_functions(db_pattern)

##
#  BEGIN DEF create_users(secret_match, commit, log_file)
##
def create_users(secret_match, commit, log_file):

    region='us-east-1'
    cur = None
    conn = None
    tunnel_open = False

    skip = True

    session = boto3.session.Session()
    client = session.client(service_name='secretsmanager', region_name=region)

    secrets_response = get_secrets_list(client)

    for secret in secrets_response:

        secret_name = secret['Name']
       
        #if secret_name.find('prod-us-east-1-pg-sm') >= 0 and secret_name.find(secret_match) >= 0:
        if 'qa-us-east-1-pg-sm' in secret_name and secret_match in secret_name and 'z-qa' not in secret_name:
            
            if 'lager' in secret_name:
                skip = False 

            if not skip:
                try:
                    response = client.get_secret_value(SecretId=secret_name)
                except (Exception) as error:
                    print('SECRET:', secret_name, file=log_file)
                    print('SECRET:', secret_name)
                    continue    

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']
                    #print('\tsecret_s:', secret_s)

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                    #print('\tsecret_b:', secret_b)
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                port = secret_dict['port']

                # ssh tunnel to aws pg db
                tunnel_cmd = OPEN_SSH_TUNNEL + host

                #print()
                #print("Exec CMD: " + tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True 

            #print()
            #print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
            
                try:
                    conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)
                    print()
                    print('DATABASE: ', db_name)
                except(psycopg.DatabaseError) as error:
                    print(error)
                    tunnel_cmd = CLOSE_SSH_TUNNEL + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
                    continue

                try:
                    cur = conn.cursor()
                    
                    sqls = CREATE_NOVUSU1_SQL
                    cur.execute(sqls)

                    sqls = CREATE_NOVUSU2_SQL
                    cur.execute(sqls)

                    sqls = CREATE_NOVUSR1_SQL
                    print(cur.execute(sqls))

                    sqls = CREATE_NOVUSR2_SQL
                    cur.execute(sqls)

                    
                    sqls = NOVUSU1_LOGIN_SQL
                    cur.execute(sqls)

                    sqls = NOVUSU2_LOGIN_SQL
                    cur.execute(sqls)

                    sqls = NOVUSR1_LOGIN_SQL
                    cur.execute(sqls)

                    sqls = NOVUSR2_LOGIN_SQL
                    cur.execute(sqls)


                    sqls = NOVUSU1_PASSWORD_SQL
                    cur.execute(sqls)

                    sqls = NOVUSU2_PASSWORD_SQL
                    cur.execute(sqls)

                    sqls = NOVUSR1_PASSWORD_SQL
                    cur.execute(sqls)

                    sqls = NOVUSR2_PASSWORD_SQL
                    cur.execute(sqls)

                    sqls = GRANT_RR1_SQL
                    cur.execute(sqls)

                    sqls = GRANT_RR2_SQL
                    cur.execute(sqls)

                    sqls = GRANT_RU1_SQL
                    cur.execute(sqls)

                    sqls = GRANT_RU2_SQL
                    cur.execute(sqls)

                    sqls = GRANT_UU2_SQL
                    cur.execute(sqls)

                    sqls = GRANT_UU1_SQL
                    cur.execute(sqls)

                    if commit:
                        conn.commit()

                    cur.close()
        
                except (Exception, psycopg.DatabaseError) as error:
                    print(error)
                    tunnel_cmd = CLOSE_SSH_TUNNEL + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
                    continue
                finally:
                    if cur is not None:
                        cur.close()
                        cur = None
                        print('Database cursor closed.')
                    if conn is not None:
                        conn.close()
                        conn = None
                        print('Database connection closed.')
                    if tunnel_open:
                        tunnel_cmd = CLOSE_SSH_TUNNEL + host
                        os.system(tunnel_cmd)
                        tunnel_open = False

    if cur is not None:
        cur.close()
        print('Database cursor closed.')
    if conn is not None:
        conn.close()
        print('Database connection closed.')
    if tunnel_open:
        tunnel_cmd = CLOSE_SSH_TUNNEL + host
        os.system(tunnel_cmd)
        tunnel_open = False

#  END DEF create_users(secret_match, commit, log_file)

##
#  BEGIN DEF generate_db_traffic()
##
def generate_db_traffic(dbname, user, password, count):
    
    conn = psycopg.connect(dbname, user, password)

    for count in range(1000):
        
        cur = conn.cursor()
        #cur.execute("SELECT * FROM tocblrmain.node;")
        insert_stmt = "INSERT INTO tocblrmain.node VALUES ('asdfasdfasdf_3002" + str(count) + "', 3002" + str(count) + ", 'N', 'asdfasdfasdf', 'qwer', 123" + str(count) + ", 'asdfasdfasdf', 'aqwegasd', 'asdferwags', 'asdfgwergas', 'aaweasd', 123000" + str(count) + ");"
        print(insert_stmt)
        cur.execute(insert_stmt)
        cur.execute("SELECT * FROM tocblrmain.node;")
        conn.commit()
        print(count,". ",cur.fetchone())
        cur.close()
        
    conn.close()

# END DEF generate_db_traffic()

##
#  BEGIN DEF audit_trigger_functions(log, commit)
##
def audit_trigger_functions(db_filter, commit):
   
    skip_schemas = [
                    'datadog',
                    'heartbeat_monitor',
                    'migration_admin',
                    'perfstat',
                    'appqossys',
                    'stdbyperf',
                    'stdbyperf',
                    'perfstat',
                    'sys',
                    'wmsys',
                    'outln',
                    'tsmsys',
                    'system',
                    'migration_admin',
                    'information_schema',
                    'public',
                    'partman', 
                    'repack',
                    'sys.whr',
                    'aws_oracle_context',
                    'aws_oracle_data',
                    'aws_oracle_ext',
                    'aws_postgis',
                    'aws',
                    'dbsnmp',
                    'killsession_user',
                    'dba_utls',
                    'aws_oracle_context'
                ]
   
    try:
    
        region='us-east-1'
        cur = None
        conn = None
        tunnel_open = False
        commit_replace = commit
        commit_drop = False
        commit = False

        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)
        
        log_file = os.getcwd() + '\\log\\trigger_audit.log'
        log = open(log_file, 'w')

        secrets_response = get_secrets_list(client)

        for secret in secrets_response:

            secret_name = secret['Name']

            #print('secret_name:', secret_name)
            print(".", end='')

            if 'prod-us-east-1-pg-sm' in secret_name and db_filter in secret_name:
                print()
                try:
                    response = client.get_secret_value(SecretId=secret_name)
                except (Exception, psycopg.DatabaseError) as error:
                    # skip duplicates
                    print()
                    print('secret_name', secret_name)
                    print(error)
                    continue

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']
                    print('\tsecret_s:', secret_s)

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                    print('\tsecret_b:', secret_b)
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                #cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                #port = secret_dict['port']

                open_tunnel = "cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I -c "
                close_tunnel ="cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I --close -c "

                # ssh tunnel to aws pg db
                tunnel_cmd = open_tunnel + host

                #print()
                print("Exec CMD: " + tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True

                #print()
                #print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
                #print()

                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                sql = "SELECT routine_schema, routine_name, routine_definition FROM information_schema.routines "
                sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t%\';"
                #sql += "WHERE routine_type = \'FUNCTION\';"

                print(sql)

                cur = conn.cursor()
                cur.execute(sql)
                conn.commit()
                schema_routines = cur.fetchall()

                #print('db_name:',db_name)
                #print('db_name:',db_name, file=log)
                
                for schema_routine in schema_routines:
                    routine_schema = schema_routine[0]
                    routine_name = schema_routine[1]
                    routine_definition = schema_routine[2]

                    if 'oldrow' in routine_definition:

                        print(db_name, routine_schema, routine_name, routine_definition)

                    # break up the delete function sql into an array
                    schema_def = routine_schema + '.'
                    schema_split = routine_definition.split(' ')

                    sql_tables = []

                    #split the table names from the schema names in the sql array items
                    for segment in schema_split:

                        segment = segment.rstrip('\n')
                        segment = segment.rstrip('\r')
                        segment = segment.lstrip('\r')
                        segment = segment.lstrip('\n')

                        if schema_def in segment:
                            table_split = segment.split('.')

                            #print(table_split)
                            #print('\t', table_split[1])

                            if table_split[1] not in sql_tables:
                                sql_tables.append(table_split[1]) #table name is always 2nd item

                    table_sql = "SELECT table_name FROM information_schema.tables WHERE table_schema=\'" + routine_schema + "\';"

                    cur.execute(table_sql)
                    conn.commit()
                    schema_tables = cur.fetchall()
                    
                    schm_tbls = []

                    #place table name str from sql response in an array
                    for table in schema_tables:
                        table_name = table[0]

                        if table_name not in schm_tbls:
                            schm_tbls.append(table_name)

                    # verify table name in delete sql exists

                    if(len(sql_tables) == 0):

                        # empty schema function, DROP
                        print()
                        print('DELETE:', routine_schema + '.' + routine_name)

                        drop_sql = 'DROP FUNCTION IF EXISTS ' + routine_schema + '.' + routine_name + ';'
                        
                        print(drop_sql)

                        try:
                            if(commit_drop):
                                cur.execute(drop_sql)
                                conn.commit()
                        except (Exception, psycopg.DatabaseError) as error:
                            # skip duplicates
                            print()
                            print(error)
                            
                            conn.rollback()
                            #cur.close()
                            continue
                    else:
                        for table in sql_tables:

                            if 'oldrow' not in routine_definition:
                                print("NO FIX:", routine_definition)
                                continue

                            #if table not in schm_tbls:
                                
                            '''
                                print()
                                print()
                                print(db_name, routine_schema, routine_name)
                                print('\tNOT FOUND:', table) 
                            

                                print(file=log)
                                print(file=log)
                                print(db_name, routine_schema, routine_name, file=log)
                                print('\tNOT FOUND:', table, file=log) 
                            '''

                            temp_routine_definition = " ".join(routine_definition.split())
                            temp_routine_definition = temp_routine_definition[5:]
                            temp_routine_definition = temp_routine_definition[:-5]
                            
                            rd_stmts = temp_routine_definition.split(';')
                            
                            new_routine_definition = "BEGIN\n    "
                            
                            for stmt in rd_stmts:
                                
                                stmt = stmt.lstrip(" ")

                                #if table not in stmt:
                                if 'RETURN' not in stmt:
                                    new_routine_definition += (stmt + ';\n    ')
                                else:
                                    new_routine_definition = new_routine_definition.lstrip(" ")
                                    new_routine_definition += stmt
                            
                            new_routine_definition = new_routine_definition.rstrip("    ")
                            new_routine_definition += 'END; '
                            
                            new_routine_definition = new_routine_definition.replace("oldrow", "old")

                            if 'DELETE' not in new_routine_definition and commit_drop:
                                #empty function, delete
                                print()
                                print('DELETE:', routine_schema + '.' + routine_name)

                                drop_sql = 'DROP FUNCTION IF EXISTS ' + routine_schema + '.' + routine_name + ';'
                                
                                print(drop_sql)

                                try:
                                    if(commit_drop):
                                        cur.execute(drop_sql)
                                        conn.commit()
                                except (Exception, psycopg.DatabaseError) as error:
                                    # skip duplicates
                                    print()
                                    print(error)
                                    
                                    conn.rollback()
                                    #cur.close()
                                    continue

                            else:
                                #replace with new routine_definition
                                print()
                                print('REPLACE:', routine_schema + '.' + routine_name)

                                replace_sql = "CREATE OR REPLACE FUNCTION " + routine_schema + "." + routine_name + "() "
                                replace_sql += " RETURNS trigger "
                                replace_sql += " LANGUAGE plpgsql "
                                replace_sql += "AS $function$ "
                                replace_sql += new_routine_definition
                                replace_sql += "$function$;"

                                print(replace_sql)

                                if(commit_replace):
                                    cur.execute(replace_sql)
                                    conn.commit()

                                print()
                                print('Old Definition:')
                                print(routine_definition)
                                print()
                                print('New Definition:')
                                print(new_routine_definition)

                                print(file=log)
                                print('Old Definition:', file=log)
                                print(routine_definition, file=log)
                                print(file=log)
                                print('New Definition:', file=log)
                                print(new_routine_definition, file=log)

                if(conn is not None):
                    conn.close()
                if(cur is not None):
                    cur.close()
                if tunnel_open:
                    tunnel_cmd = close_tunnel + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
    
    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)

    finally:

        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()
        if tunnel_open:
                    tunnel_cmd = close_tunnel + host
                    os.system(tunnel_cmd)
                    tunnel_open = False

#  END DEF audit_trigger_functions(log, commit)

##
#  BEGIN DEF update_trigger_functions(db_filter, commit)
##
def update_trigger_functions(db_filter, commit):
    
    skip_schemas = [
                    'datadog',
                    'heartbeat_monitor',
                    'migration_admin',
                    'perfstat',
                    'appqossys',
                    'stdbyperf',
                    'stdbyperf',
                    'perfstat',
                    'sys',
                    'wmsys',
                    'outln',
                    'tsmsys',
                    'system',
                    'migration_admin',
                    'information_schema',
                    'public',
                    'partman',
                    'repack',
                    'sys.whr',
                    'aws_oracle_context',
                    'aws_oracle_data',
                    'aws_oracle_ext',
                    'aws_postgis',
                    'aws',
                    'dbsnmp',
                    'killsession_user',
                    'dba_utls',
                    'aws_oracle_context'
                ]
    region='us-east-1'
    cur = None
    conn = None
    tunnel_open = False
    commit_replace = False
    commit_drop = False

    try:
    
        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)
        
        log_file = os.getcwd() + '\\log\\trigger_audit_nims.log'
        log = open(log_file, 'w')

        secrets_response = get_secrets_list(client)

        skip_secret = True

        for secret in secrets_response:
            skip_secret = True
            secret_name = secret['Name']

            print(".", end='')

            # pick up where we left off when the credentials inevitably time out. we can only do things that don't require much effort because security
            if 'nims' in secret_name:
                skip_secret = False

            if skip_secret is not True and 'prod-us-east-1-pg-sm' in secret_name and db_filter in secret_name and 'z-prod' not in secret_name:
                print()
                try:
                    response = client.get_secret_value(SecretId=secret_name)
                except (Exception, psycopg.DatabaseError) as error:
                    # skip duplicates
                    print()
                    print('secret_name', secret_name)
                    print(error)
                    continue

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']
                    print('\tsecret_s:', secret_s)

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                    print('\tsecret_b:', secret_b)
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                #cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                #port = secret_dict['port']

                open_tunnel = "cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I -c "
                close_tunnel ="cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I --close -c "

                # ssh tunnel to aws pg db
                tunnel_cmd = open_tunnel + host

                #print()
                print("Exec CMD: " + tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True

                #print()
                #print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
                #print()

                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                sql = "SELECT routine_schema, routine_name, routine_definition FROM information_schema.routines "
                sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t%\';"
                #sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t%\' AND routine_schema LIKE \'%wtoc%\';"
                #sql += "WHERE routine_type = \'FUNCTION\';"

                print(sql)

                cur = conn.cursor()
                cur.execute(sql)
                conn.commit()
                schema_routines = cur.fetchall()

                #print('db_name:',db_name)
                #print('db_name:',db_name, file=log)
                
                for schema_routine in schema_routines:
                    routine_schema = schema_routine[0]
                    routine_name = schema_routine[1]
                    routine_definition = schema_routine[2]

                    if routine_schema in skip_schemas:
                        break

                    # identify and correct references to on prem 'oldrow'
                    if 'oldrow' in routine_definition:
                        
                        new_routine_definition = routine_definition.replace('oldrow', 'old')

                        #print(db_name, routine_schema, routine_name, new_routine_definition)

                        replace_sql = "CREATE OR REPLACE FUNCTION " + routine_schema + "." + routine_name + "() "
                        replace_sql += "\n   RETURNS trigger "
                        replace_sql += "LANGUAGE plpgsql "
                        replace_sql += "AS $function$ "
                        replace_sql += new_routine_definition
                        replace_sql += "$function$;"

                        print()
                        print(replace_sql)

                        if(commit):
                            cur.execute(replace_sql)
                            conn.commit()

                    # break up the delete function sql into an array
                    schema_def = routine_schema + '.'
                    schema_split = routine_definition.split(' ')

                    sql_tables = []

                    #split the table names from the schema names in the sql array items
                    for segment in schema_split:

                        segment = segment.rstrip('\n')
                        segment = segment.rstrip('\r')
                        segment = segment.lstrip('\r')
                        segment = segment.lstrip('\n')

                        if schema_def in segment:
                            table_split = segment.split('.')

                            #print(table_split)
                            #print('\t', table_split[1])

                            if table_split[1] not in sql_tables:
                                sql_tables.append(table_split[1]) #table name is always 2nd item

                    table_sql = "SELECT table_name FROM information_schema.tables WHERE table_schema=\'" + routine_schema + "\';"

                    cur.execute(table_sql)
                    conn.commit()
                    schema_tables = cur.fetchall()
                    
                    schm_tbls = []
                    
                    print()
                    print(routine_schema)

                    #place table name str from sql response in an array
                    for table in schema_tables:
                        table_name = table[0]
                        found_cid = False
                        
                        print('\t', table_name)

                        if table_name not in schm_tbls:
                            schm_tbls.append(table_name)

                        col_sql  = "SELECT * FROM information_schema.columns "
                        col_sql += "WHERE table_schema = \'" + routine_schema + "\' "
                        col_sql += "AND table_name = \'" + table_name + "\';"

                        cur.execute(col_sql)
                        conn.commit()  
                        columns = cur.fetchall()
                        
                        for column in columns:
                            #print('\t\t', column[3])

                            if 'collection_id' in column[3]:
                               found_cid = True
                               #print('\t\tFOUND CID FOR:', table_name)
                               break
                        
                        if not found_cid:
                            #print('\t\tNO CID FOR:', table_name)

                            if table_name not in routine_definition:
                                print('\t\t', table_name, ' NOT FOUND in ', routine_schema)

                                if 'xml_equivalency' not in table_name:

                                    pos = routine_definition.find('RETURN')

                                    temp_definition = routine_definition[:pos - 1]
                                    temp_definition += "DELETE FROM " + routine_schema + '.' + table_name + " \n\tWHERE old.root_id = "
                                    temp_definition += routine_schema + '.' + table_name + '.root_id;'
                                    temp_definition += "\nRETURN OLD;\nEND;"  

                                    replace_sql = "CREATE OR REPLACE FUNCTION " + routine_schema + "." + routine_name + "() "
                                    replace_sql += "\n   RETURNS trigger "
                                    replace_sql += "LANGUAGE plpgsql "
                                    replace_sql += "AS $function$ "
                                    replace_sql += temp_definition
                                    replace_sql += "\n$function$;"

                                    print()
                                    print(replace_sql)

                                    if(commit):
                                        cur.execute(replace_sql)
                                        conn.commit()

                            #check for missing functions and add as needed
                            func_sel_sql = "SELECT * FROM information_schema.triggers WHERE trigger_schema = \'" + routine_schema + "\'"

                            #print()
                            #print(routine_schema)

                            cur.execute(func_sel_sql)
                            conn.commit()  
                            functions = cur.fetchall()
                            
                            if len(functions) > 0:

                                found = False
                                for function in functions:
                                    trigger_name = function[2]

                                    if 'sidetab' not in trigger_name:
                                        print(routine_schema + ':', function)
                                    else:
                                        found = True
                                    
                                    if not found:
                                        print('NOT FOUND sidetab:', db_name, routine_schema)

                                        create_sql = "CREATE TRIGGER sidetab_del_t "
                                        create_sql += "\n   AFTER DELETE ON  " + routine_schema + ".nims_payload  "
                                        create_sql += "\n      FOR EACH ROW EXECUTE FUNCTION " + routine_schema + ".\"sidetab_del_t$nims_payload\"(); "

                                        print(create_sql)

                                        if(commit):
                                            cur.execute(create_sql)
                                            conn.commit()
                                    #else:
                                        #print("\t\tTrigger function found:", trigger_name)
                                        #print()
                                        #print('\t\t', function)

                            else:
                                print('NOT FOUND trigger:', db_name, routine_schema)

                                create_sql = "CREATE TRIGGER sidetab_del_t "
                                create_sql += "\n   AFTER DELETE ON  " + routine_schema + ".nims_payload  "
                                create_sql += "\n      FOR EACH ROW EXECUTE FUNCTION " + routine_schema + ".\"sidetab_del_t$nims_payload\"(); "

                                print(create_sql)

                                if(commit):
                                    cur.execute(create_sql)
                                    conn.commit()

                if(conn is not None):
                    conn.close()
                if(cur is not None):
                    cur.close()
                if tunnel_open:
                    tunnel_cmd = close_tunnel + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
    
    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)

    finally:

        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()
        if tunnel_open:
            tunnel_cmd = close_tunnel + host
            os.system(tunnel_cmd)
            tunnel_open = False

#  END DEF update_trigger_functions(log, commit)

##
# BEGIN DEF populate_migration_table
# populate_migration_table()
##
def populate_migration_table(commit):

    target_resource = 'lala'
    source_resource = 'lulu'
    source_region = 'EGN'
    target_region = 'USE1'

    db_name = 'pgloadqueue01ap'
    db_user = 'postgres'
    db_password = ';-!.3,R}1,nJ,O,h}j}=|`EE`y]T7$`J'
    db_password = '%Tmx^*FF7z}j|+au]BZ6NGa8l+01eszo'

    #filename = "qc_doc_col_map.xlsx"
    #wb = openpyxl.load_workbook(filename)
    #ws = wb.active

    conn = None
    cur = None

    try:
        conn = psycopg.connect(dbname=db_name, user=db_user, password=db_password, host="localhost", port="8383")
            
        cci_name =  'DEFAULT'
        source_region =  'EGN'
        content_type = None 
        source_resource = None
        target_region =  'USE1'
        target_resource =  None

        rsc_types = migmaps.PROD_DB_TYPE_DICT
        op_rsc_map = migmaps.PROD_MASTER_RSC_LIST_NI

        for op_rsc in op_rsc_map:
            source_resource = op_rsc
            target_resource = op_rsc_map[op_rsc]
            content_type = rsc_types[target_resource]

            #print(source_resource, target_resource, content_type)

            #source_resource = list(migmaps.PROD_MASTER_RSC_LIST.keys())[list(migmaps.PROD_MASTER_RSC_LIST.values()).index(target_resource)]
            #source_resource = migmaps.PROD_MASTER_RSC_LIST.keys()[migmaps.PROD_MASTER_RSC_LIST.values().index(target_resource)] 

            print()
            print('source_resource:', source_resource, ' : ', 'target_resource:', target_resource, 'content_type:', content_type)

            cur = conn.cursor()

            sqls  = "SELECT content_type, cci_name, source_region, source_resource, target_region, target_resource, last_upd_time "
            sqls += "FROM load.region_resource_mapping "
            sqls += "WHERE content_type=\'" + content_type 
            sqls += "\' AND source_resource =\'" + source_resource 
            sqls += "\' AND target_resource=\'" + target_resource
            sqls += "\' AND cci_name=\'" + cci_name
            sqls += "\' AND source_region=\'" + source_region
            sqls += "\' AND target_region=\'" + target_region + "';"
    
            try:
                
                cur.execute(sqls)
                conn.commit

                maps = cur.fetchall()
                
                print("row_count:", cur.rowcount)
                length = len(maps)

                print('LENGTH', length)
            
                if len(maps) < 1:
                    sqli = "INSERT INTO load.region_resource_mapping (content_type, cci_name, source_region, source_resource, target_region, target_resource, last_upd_time) VALUES (" 
                    sqli += "\'" + content_type + "\', \'" + cci_name + "\', \'" + source_region + "\', \'" + source_resource + "\', \'"  + target_region + "\', \'"  + target_resource + "\', " 
                    sqli += "current_timestamp(0));"
                    
                    print(sqli)
                    
                    try:
                        if(commit):
                            cur.execute(sqli) 
                            conn.commit()
                    except (Exception, psycopg.DatabaseError) as error:
                        print()
                        print(error)  
                        print(sqls)
                        conn.rollback()

                    if(content_type == 'DOC'):
                        
                        content_type = 'TOC'
                        
                        sqli = "INSERT INTO load.region_resource_mapping (content_type, cci_name, source_region, source_resource, target_region, target_resource, last_upd_time) VALUES (" 
                        sqli += "\'" + content_type + "\', \'" + cci_name + "\', \'" + source_region + "\', \'" + source_resource + "\', \'"  + target_region + "\', \'"  + target_resource + "\', " 
                        sqli += "current_timestamp(0));"

                        print(sqli)

                        try:
                            if(commit):
                                cur.execute(sqli) 
                                conn.commit()
                        except (Exception, psycopg.DatabaseError) as error:
                            print()
                            print(error)  
                            print(sqls)
                            conn.rollback()

            except (Exception, psycopg.DatabaseError) as error:
                print()
                print(error)  
                print(sqls)

    finally:
        
        if(conn):
            conn.close()
        if(cur):
            cur.close()

def test_cx_oracle():
    connection = cx_Oracle.connect(user="novusu1", password="novusu1",
                               dsn="jdbc:oracle:thin:@(description=(enable=broken)(address_list=(load_balance=no)(address=(protocol=tcp)(port=1521)(host=c153sjdpndbcci.int.thomsonreuters.com)))(connect_data=(service_name=rw_novuscore_prod.int.thomsonreuters.com)))")

    print(connection)

SKIP_SCHEMAS = [
                    'datadog',
                    'audsys',
                    'heartbeat_monitor',
                    'migration_admin',
                    'perfstat',
                    'appqossys',
                    'stdbyperf',
                    'stdbyperf',
                    'perfstat',
                    'sys',
                    'wmsys',
                    'outln',
                    'tsmsys',
                    'system',
                    'migration_admin',
                    'information_schema',
                    'public',
                    'partman',
                    'repack',
                    'sys.whr',
                    'aws_oracle_context',
                    'aws_oracle_data',
                    'aws_oracle_ext',
                    'aws_postgis',
                    'aws',
                    'dbsnmp',
                    'killsession_user',
                    'dba_utls',
                    'aws_oracle_context'
                ]

##
#  BEGIN DEF update_trigger_functions(db_filter, commit)
##
def print_db_tables(db_filter):
    
    region='us-east-1'
    cur = None
    conn = None
    tunnel_open = False
    commit_replace = False
    commit_drop = False

    try:
        
        log_file = os.getcwd() + "\\log\\prod_xml_eq_audit.log"
        log = open(log_file, 'w')

        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)
        
        log_file = os.getcwd() + '\\log\\trigger_audit.log'
        log = open(log_file, 'w')

        secrets_response = get_secrets_list(client)

        skip_secret = True

        for secret in secrets_response:

            secret_name = secret['Name']

            print(".", end='')

            # pick up where we left off when the credentials inevitably time out. we can only do things that don't require much effort because security
            if 'prod' in secret_name: # super secret double probation override. Kludge ftw.
                skip_secret = False

            if skip_secret is not True and 'prod-us-east-1-pg-sm' in secret_name and db_filter in secret_name and 'z-prod' not in secret_name:
                print()
                try:
                    response = client.get_secret_value(SecretId=secret_name)
                except (Exception, psycopg.DatabaseError) as error:
                    # skip duplicates
                    print()
                    print('secret_name', secret_name)
                    print(error)
                    continue

                # Decrypts secret using the associated KMS CMK.
                # Depending on whether the secret is a string or binary, one of these fields will be populated.
                if 'SecretString' in response:
                    secret_s = response['SecretString']
                    print('\tsecret_s:', secret_s)

                elif 'SecretBinary' in response:
                    secret_b = base64.b64decode(response['SecretBinary'])
                    print('\tsecret_b:', secret_b)
                else:
                    print('\tSecret String NOT FOUND')

                secret_dict = json.loads(secret_s) 

                #cluster_id = secret_dict['dbClusterIdentifier'] 
                host = secret_dict['host']
                db_name = secret_dict['dbname']
                user = secret_dict['username']
                password = secret_dict['password']
                #port = secret_dict['port']

                open_tunnel = "cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I -c "
                close_tunnel ="cloud-tool ssh-tunnel -r " + LOCAL_PORT + " -I --close -c "

                # ssh tunnel to aws pg db
                tunnel_cmd = open_tunnel + host

                #print()
                print("Exec CMD: " + tunnel_cmd)
                os.system(tunnel_cmd)
                tunnel_open = True

                #print()
                #print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')
                #print()

                conn = psycopg2.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)

                #schema_names = get_schema_names(conn)
                #print(schema_names)

                tables = get_tables(conn)

                for row in tables:
                    if 'xml_equiv' in row["table_name"] and row["table_schema"] not in SKIP_SCHEMAS:
                        print("{}.{}".format(db_name, row["table_schema"]))
                        print("{}.{}".format(db_name, row["table_schema"]), file=log)

                sql = "SELECT routine_schema, routine_name, routine_definition FROM information_schema.routines "
                sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t%\';"
                #sql += "WHERE routine_type = \'FUNCTION\' AND routine_name LIKE \'%sidetab_del_t%\' AND routine_schema LIKE \'%wtoc%\';"
                #sql += "WHERE routine_type = \'FUNCTION\';"

                print(sql)

                cur = conn.cursor()
                cur.execute(sql)
                conn.commit()
                schema_routines = cur.fetchall()

                #print('db_name:',db_name)
                #print('db_name:',db_name, file=log)
                
                for schema_routine in schema_routines:
                    routine_schema = schema_routine[0]
                    routine_name = schema_routine[1]
                    routine_definition = schema_routine[2]


                if(conn is not None):
                    conn.close()
                if(cur is not None):
                    cur.close()
                if tunnel_open:
                    tunnel_cmd = close_tunnel + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
    
    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)

    finally:

        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()
        if tunnel_open:
            tunnel_cmd = close_tunnel + host
            os.system(tunnel_cmd)
            tunnel_open = False

#  END DEF update_trigger_functions(log, commit)

#
## BEGIN get_schema_names(connection)
#
def get_schema_names(connection):
        
        sql = """SELECT nspname FROM pg_namespace ORDER BY nspname"""
        rp = connection.execute(sql)

        if util.py2k:
            schema_names = [row[0].decode(utf_8) for row in rp
                            if not row[0].startswith('pg_')]
        else:
            schema_names = [row[0] for row in rp
                            if not row[0].startswith('pg_')]
        
        print(schema_names)
        
        return schema_names 

## END get_schema_names(connection)

#
## BEGIN get_tables(connection)
#
def get_tables(connection):

    """
    Create and return a list of dictionaries with the
    schemas and names of tables in the database
    connected to by the connection argument.
    """
    cursor_factory = None
    cursor = connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cursor.execute("""SELECT table_schema, table_name
                      FROM information_schema.tables
                      WHERE table_schema != 'pg_catalog'
                      AND table_schema != 'information_schema'
                      AND table_type='BASE TABLE'
                      ORDER BY table_schema, table_name""")

    tables = cursor.fetchall()

    cursor.close()

    return tables

## END get_tables(connection)

def print_tables(tables):

    """
    Prints the list created by get_tables
    """

    for row in tables:

        print("{}.{}".format(row["table_schema"], row["table_name"]))

    return tables

##
#  BEGIN DEF create_doc_admin_role()
##
def create_doc_admin_user():
    try:
        
        conn = None
        cur = None

        region='us-east-1'
        cur = None
        conn = None
        tunnel_open = False

        session = boto3.session.Session()
        client = session.client(service_name='secretsmanager', region_name=region)
        
        secrets_response = get_secrets_list(client)

        found = False
        inner = False

        for secret in secrets_response:

            secret_name = secret['Name']

            #if (''a207947-doc-westlaw-04a-qa-us-east-1-pg-sm'' in secret_name) or inner:
            if ('a207947-doc-news-03a-qa-us-east-1-pg-sm' in secret_name) or inner:
                
                inner = True
                
                if 'qa-us-east-1-pg-sm' in secret_name and \
                    'doc' in secret_name and \
                    'docloc' not in secret_name and \
                    'dlc' not in secret_name:

                    response = client.get_secret_value(SecretId=secret_name)

                    # Decrypts secret using the associated KMS CMK.
                    # Depending on whether the secret is a string or binary, one of these fields will be populated.
                    if 'SecretString' in response:
                        secret_s = response['SecretString']
                        print('\tsecret_s:', secret_s)

                    elif 'SecretBinary' in response:
                        secret_b = base64.b64decode(response['SecretBinary'])
                        print('\tsecret_b:', secret_b)
                    else:
                        print('\tSecret String NOT FOUND')
                        continue

                    secret_dict = json.loads(secret_s) 

                    host = secret_dict['host']
                    db_name = secret_dict['dbname']
                    user = secret_dict['username']
                    password = secret_dict['password']

                    # ssh tunnel to aws pg db
                    tunnel_cmd = OPEN_SSH_TUNNEL + host

                    try:

                        print()
                        print("Exec CMD: " + tunnel_cmd)
                        os.system(tunnel_cmd)
                        tunnel_open = True

                        print()
                        print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost, 8899...')
                        print()

                        conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=8899)
                
                        sql = get_add_db_user_sql('novusdocadmin', 'asadmin')
                        cur = conn.cursor()
                        cur.execute(sql)
                        cur.execute("GRANT postgres TO novusdocadmin")
                        conn.commit()
                        
                        grant_read_access('novusdocadmin', conn)

                        schema_sql = """SELECT schema_name FROM information_schema.schemata WHERE schema_name NOT LIKE \'pg_%\' AND schema_name NOT LIKE \'aws%\';"""
                        #schema_sql = """SELECT schema_name FROM information_schema.schemata WHERE schema_name = \'pglogical\';"""

                        cur = conn.cursor()
                        cur.execute(schema_sql)
                        conn.commit()
                        schemas = cur.fetchall()

                        for schema in schemas:  
                
                            schema = schema[0]

                            if schema in SKIP_SCHEMAS:
                                print('SKIPPING:', schema)
                            
                            else:
                                print('PROCESS:', schema)
                                try:
                                    cur = conn.cursor()
                                    
                                    cur.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = \'" + schema + "\'")
                                    #cur.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = \'" + schema + "\' AND table_name = \'compression\'") ### MAJOR KLUDGE ###

                                    conn.commit()
                                    tables = cur.fetchall()

                                    for table in tables: 
                                        
                                        sql = "GRANT TRUNCATE ON " + schema + ".\"" + table[0] + "\" TO novusdocadmin"
                                        #sql = "GRANT TRUNCATE ON " + schema + ".compression TO novusdocadmin"
                                        print(sql)
                                        cur = conn.cursor()
                                        cur.execute(sql)
                                        conn.commit()

                                except (Exception, psycopg.DatabaseError) as error:
                                    # skip not founds
                                    print()
                                    print(error) 
                                    
                                    if(cur is not None):
                                        cur.close()
                                    if(conn is not None):
                                        conn.close()
                        
                        tunnel_cmd = CLOSE_SSH_TUNNEL + host
                        os.system(tunnel_cmd)
                        tunnel_open = False
                    
                    except (Exception, psycopg.DatabaseError) as error:
                            # skip not founds
                            print()
                            print(error)
                            
                            if(cur is not None):
                                cur.close()
                            if(conn is not None):
                                conn.close()
                            if(tunnel_open):
                                tunnel_cmd = CLOSE_SSH_TUNNEL + host
                                os.system(tunnel_cmd)
                                tunnel_open = False 
            else:
                print('skipping: ' + secret_name) 

    except (Exception, psycopg.DatabaseError) as error:
        # skip duplicates
        print()
        print(error)  

    finally:

        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()
        if(tunnel_open):
            tunnel_cmd = CLOSE_SSH_TUNNEL + host
            os.system(tunnel_cmd)
            tunnel_open = False 

def timestamp_hack(commit):

    conn = None
    cur = None

    try:    
        conn = psycopg.connect(dbname='pgloadqueue01ap', user='postgres', password='.Bx9%NV7$JvP.aco&mm{[gLfRQ=<9(M?', host="localhost", port="5151")
            
        cci_name =  'DEFAULT'
        source_region =  'EGN'
        content_type = None 
        source_resource = None
        target_region =  'USE1'
        target_resource =  None

        print()
        print('source_resource:', source_resource, ' : ', 'target_resource:', target_resource, 'content_type:', content_type)

        cur = conn.cursor()

        sqls  = "SELECT content_type, cci_name, source_region, source_resource, target_region, target_resource, last_upd_time "
        sqls += "FROM load.region_resource_mapping "
        
        '''
        sqls += "WHERE content_type=\'" + content_type 
        sqls += "\' AND source_resource =\'" + source_resource 
        sqls += "\' AND target_resource=\'" + target_resource
        sqls += "\' AND cci_name=\'" + cci_name
        sqls += "\' AND source_region=\'" + source_region
        sqls += "\' AND target_region=\'" + target_region + "';"
        '''

        try:
            
            cur.execute(sqls)
            conn.commit

            maps = cur.fetchall()
            
            print("row_count:", cur.rowcount)
            length = len(maps)

            print('LENGTH', length)

            for map in maps:

                usql = "UPDATE \"load\".region_resource_mapping "
                usql += "SET last_upd_time=(current_timestamp) "
                usql += "WHERE source_resource=\'" + map[3] + "\';"

                print(usql)

                cur = conn.cursor()
                cur.execute(usql)
                
                if commit:
                    conn.commit()

        except (Exception, psycopg.DatabaseError) as error:
            print()
            print(error)  
            print(sqls)

    finally:
        
        if(conn):
            conn.close()
        if(cur):
            cur.close()


##
#  BEGIN DEF grant_access_oneoff()
#  ssh tunnel to db/port required
##
def grant_access_oneoff(dbname, password, port):

    try:
        #print()
        #print('Connecting to the PostgreSQL database ', 'pgcci01aq', 'postgres', '1H.9,<oN9mOfq<2|TZua62,W]N(Yc7_C', 'localhost', 9992, '...')
        #print()

        #conn = psycopg.connect(dbname='pgcci01aq', user='postgres', password='1H.9,<oN9mOfq<2|TZua62,W]N(Yc7_C', host="localhost", port='8891')
        conn = psycopg.connect(dbname=dbname, user='postgres', password=password, host="localhost", port=port)

        try:
            sql = get_add_db_user_sql('novusu1', 'novusu1')

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()

            sql = "ALTER USER novusu1 WITH PASSWORD \'novusu1\'; ALTER ROLE \"novusu1\" WITH LOGIN;"

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()

            sql = "ALTER USER novusr1 WITH PASSWORD \'novusr1\';  ALTER ROLE \"novusr1\" WITH LOGIN;"

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()

            sql = "ALTER USER novusr2 WITH PASSWORD \'novusr2\'; ALTER ROLE \"novusr2\" WITH LOGIN;"

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()

            sql = "ALTER USER novusu2 WITH PASSWORD \'novusu2\';  ALTER ROLE \"novusu2\" WITH LOGIN;"

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()
            
            grant_update_access('novusu1', conn)
        
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error) 
            
            if(cur is not None):
                cur.close()
            if(conn is not None):
                conn.close()
        
        try:
            sql = get_add_db_user_sql('novusu2', 'novusu2')

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()

            grant_update_access('novusu2', conn)
            #log_file = open('client_schema_audit.1209.txt', 'r')
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error)
            
            if(cur is not None):
                cur.close()
            if(conn is not None):
                conn.close()

        try:
            sql = get_add_db_user_sql('novusr1', 'novusr1')

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()

            grant_read_access('novusr1', conn)
            #log_file = open('client_schema_audit.1209.txt', 'r')
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error)
            
            if(cur is not None):
                cur.close()
            if(conn is not None):
                conn.close()

        try:
            sql = get_add_db_user_sql('novusr2', 'novusr2')

            print('SQL:', sql)
            cur = conn.cursor()
            result = cur.execute(sql)
            conn.commit()

            grant_read_access('novusr2', conn)
            #log_file = open('client_schema_audit.1209.txt', 'r')
        
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error) 

        try:
            grant_update_access('novusu1', conn)
            #log_file = open('client_schema_audit.1209.txt', 'r')
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error) 
            #continue
        
        try:
            grant_update_access('novusu2', conn)
            #log_file = open('client_schema_audit.1209.txt', 'r')
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error)

        try:
            grant_read_access('novusr1', conn)
            #log_file = open('client_schema_audit.1209.txt', 'r')
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error)

        try:
            grant_read_access('novusr2', conn)
            #log_file = open('client_schema_audit.1209.txt', 'r')
        except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error)
                    
        if(conn is not None):
            conn.close()
  
    except (Exception, psycopg.DatabaseError) as error:
            # skip not founds
            print()
            print(error)

##
#   BEGIN DEF map_schema_locations(log)
##
def map_schema_locations(service):

    region='us-east-1'
    cur = None
    conn = None
    tunnel_open = False

    session = boto3.session.Session()
    client = session.client(service_name='secretsmanager', region_name=region)

    secrets_response = get_secrets_list(client)

    for secret in secrets_response:

        secret_name = secret['Name']

        if 'prod-us-east-1-pg-sm' in secret_name and service.lower() in secret_name.lower():

            response = client.get_secret_value(SecretId=secret_name)

            # Decrypts secret using the associated KMS CMK.
            # Depending on whether the secret is a string or binary, one of these fields will be populated.
            if 'SecretString' in response:
                secret_s = response['SecretString']
                print('\tsecret_s:', secret_s)

            elif 'SecretBinary' in response:
                secret_b = base64.b64decode(response['SecretBinary'])
                print('\tsecret_b:', secret_b)
            else:
                print('\tSecret String NOT FOUND')

            secret_dict = json.loads(secret_s) 

            cluster_id = secret_dict['dbClusterIdentifier'] 
            host = secret_dict['host']
            db_name = secret_dict['dbname']
            user = secret_dict['username']
            password = secret_dict['password']
            port = secret_dict['port']

            # ssh tunnel to aws pg db
            tunnel_cmd = OPEN_SSH_TUNNEL + host

            print()
            print("Exec CMD: " + tunnel_cmd)
            os.system(tunnel_cmd)
            tunnel_open = True 

            print()
            print('Connecting to the PostgreSQL database ', db_name, user, password, 'localhost', LOCAL_PORT, '...')

            try:
                conn = psycopg.connect(dbname=db_name, user=user, password=password, host="localhost", port=LOCAL_PORT)
            except(psycopg.DatabaseError) as error:
                print(error)
                tunnel_cmd = CLOSE_SSH_TUNNEL + host
                os.system(tunnel_cmd)
                tunnel_open = False
                continue

            try: 
                cur = conn.cursor()
                cur.execute("SELECT schema_name FROM information_schema.schemata WHERE " +
                            "(schema_name NOT LIKE \'pg_%\' " +
                            "AND schema_name NOT LIKE \'repack%\' " +
                            "AND schema_name NOT LIKE \'partman%\' " +
                            "AND schema_name NOT LIKE \'public%\' " +
                            "AND schema_name NOT LIKE \'information_schema%\' " +
                            "AND schema_name NOT LIKE \'aws_%\')")

                print()
                schemas = cur.fetchall()

                for schema in schemas:
                    
                    schema_name = schema[0]
                    
                    if 'anznzlinxjudg' in schema_name.lower() or 'w_filings_metadoc' in schema_name.lower() or 'anznzlinxjudg' in schema_name.lower():

                        print()
                        print('=======================FOUND======================')
                        print('=================' + db_name + ' ' + schema_name)  
                        print('==================================================')
                        print()

            except (Exception, psycopg.DatabaseError) as error:
                print(error)
                tunnel_cmd = CLOSE_SSH_TUNNEL + host
                os.system(tunnel_cmd)
                tunnel_open = False
                continue
            finally:
                if cur is not None:
                    cur.close()
                    cur = None
                    print('Database cursor closed.')
                if conn is not None:
                    conn.close()
                    conn = None
                    print('Database connection closed.')
                if tunnel_open:
                    tunnel_cmd = CLOSE_SSH_TUNNEL + host
                    os.system(tunnel_cmd)
                    tunnel_open = False
    
    if cur is not None:
        cur.close()
        print('Database cursor closed.')
    if conn is not None:
        conn.close()
        print('Database connection closed.')
    if tunnel_open:
        tunnel_cmd = CLOSE_SSH_TUNNEL + host
        os.system(tunnel_cmd)
        tunnel_open = False

#   END DEF map_schema_locations(log)

###
# BEGIN DEF update_cci()
###
def get_collection_names_by_db(dbname, pw, port):

    sqls = None
    schema_name = None

    try:

        conn2 = psycopg.connect(dbname='pgcci01ap', user='postgres', password='aR6;M2,WTu!xqdD31![2D6cx#H}T?Y<V', host="localhost", port="1212")
        cur2 = conn2.cursor()

    except (Exception, psycopg.DatabaseError) as error:
        print()
        print(error)  
        print(sqls)

    try:

        conn = psycopg.connect(dbname=dbname, user='postgres', password=pw, host="localhost", port=port)
        cur = conn.cursor()

        sqls = 'SELECT schema_name FROM information_schema.schemata WHERE schema_name NOT LIKE \'pg%\';'
        
        print(sqls)
        cur.execute(sqls)

        schemas = cur.fetchall()

        for schema in schemas:
            schema_name = schema[0]
            #print('schema: ' + schema_name) 
            
            try:

                sqls = 'SELECT collection_name FROM cci.doc_update_process WHERE doc_schema_name = \'' + schema_name + '\';'

                cur2.execute(sqls)

                collection_names = cur2.fetchall()

                for collection_name in collection_names:
                    print(collection_name[0])   
                    #print(schema_name + ',' + collection_name[0])   

            except (Exception, psycopg.DatabaseError) as error:
                print()
                print(error)  
                print(sqls)
    
    except (Exception, psycopg.DatabaseError) as error:
            print()
            print(error)  
            print(sqls)

    finally:
            
        if(conn is not None):
            conn.close()
        if(cur is not None):
            cur.close()

        if(conn2 is not None):
            conn2.close()
        if(cur2 is not None):
            cur2.close()

# END DEF update_cci()

def disable_dl():

    db_name = 'pgcci01aq'
    user = 'postgres'
    password = '=L3jO&K8XBG~MAr;X-0g}mR6G*Wj&)h#'

    collection = ''

    sql = 'SELECT collection_name, region_name, is_master, is_active, last_upd_time FROM \"load\".region_control;'
    #sql2 = 'UPDATE \"load\".region_control SET is_active=\'N\', last_upd_time=current_timestamp  WHERE collection_name=\'D-ALBAWABATR-T001\' AND region_name=\'USE1\';'
    #sql3 = 'UPDATE \"load\".region_control SET is_active=\'N\', last_upd_time=current_timestamp  WHERE collection_name=\'D-ALBAWABATR-T001\' AND region_name=\'EGN\';'

    conn = psycopg2.connect(dbname=db_name, user=user, password=password, host="localhost", port='1616')

    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    #regions = cur.fetchall()

    #print('db_name:',db_name)
    #print('db_name:',db_name, file=log)

    #for region in regions:
    #    print(region)

def update_res():

    db_name = 'pgcci01aq'
    user = 'postgres'
    password = '7i<iy=aT2ilbf3%`7SrElyMbo`Rj9VJi'

    collection = ''

    sql = 'SELECT collection_name, doc_elem_set_name, doc_resource_name, doc_schema_name FROM cci.doc_update_process';
    sql2 = 'SELECT collection_name, stage_id, doc_resource_name, doc_schema_name FROM cci.doc;'

    conn = psycopg2.connect(dbname=db_name, user=user, password=password, host="localhost", port='1515')

    cur = conn.cursor()
    cur.execute(sql2)
    conn.commit()
    rows = cur.fetchall()

    for row in rows:
        sql3 = 'UPDATE "cci".doc_update_process SET doc_resource_name=\'' + row[2] + '\' WHERE  collection_name=\'' + row[0] + '\'' 
        print(sql3)
        cur.execute(sql3)
        conn.commit()

'''
THIS IS BAD. REMOVE. call functions from awsutil.py 
'''
#grant_access_oneoff('pgnortshared01ap', 'N..9E&6x]Q$_%W&k34eornu,*kfl<+MF', '7878')
grant_user_access('doc', 'doc-tax-01a-qa-us-east-1-pg-sm')

#get_collection_names_by_db('pgdocwestlaw02ap', 'SL2ATaZUL[fiRC(y}&>Mh|5VO_%c!2jZ', '1111')
#get_collection_names_by_db('pgdocwestlaw10ap', '<[tt|mFvb^S~dM;ZdRg#e,PP_*l|k2Ar', '1313')
#get_collection_names_by_db('pgdocwestlaw11ap', 'i1z,tTDn`xy%B%~[S6V^~AXCmTw2CqKA', '1414')

#create_doc_admin_user()


#disable_dl()

#update_res()

#grant_access_oneoff('pgdocwestlaw01aq', 'mS`+*Hdi2VAm*1jVw]MYQy9-juP_m8XA', '2222')
#grant_access_oneoff('pgdocwestlaw02aq', 'M![Q5aKP5$vU%-71DP9m<5<Z)x$b|iJY', '2424')

#audit_trigger_functions('norm-shared', True)

#map_schema_locations('doc')

#timestamp_hack(True)

#conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8892")
        #conn = psycopg.connect(dbname='pgcci01bi', user='postgres', password='|T&TGQMu0UM]R][fG`mGdz^n6<5hFJ<G', host="localhost", port="8891")
        #conn = psycopg.connect(dbname='pgcci01aq', user='postgres', password='IeIkr-KBQV}L{Ka<!}>Y,5$g4-y%Wc~u', host="localhost", port="8899")
        #conn = psycopg.connect(dbname='pgcci01aq', user='novusu2', password='novusu2', host="localhost", port="8891")
        #conn = psycopg.connect(dbname='pgcci01bq', user='novusu2', password='novusu2', host="localhost", port="8894")
        #conn = psycopg.connect(dbname='pgnimsccscaleretr01ai', user='postgres', password='}yEnG#$}ZeJwo>[xg)59,9?HW{pqF`1v', host="localhost", port="8891")