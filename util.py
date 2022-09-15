__author__ = "Jeremy Leese | jeremy.leese@thomsonreuters.com | Technology - Product Eng - Content Platforms"
__copyright__ = "Copyright© (C) 2022 Thomson Reuters. All Rights Reserved."
__version__ = "1.0"

import boto3
import openpyxl
import sys

USE_REGION="us-east-1"
SM_SERVICE="secretsmanager"
RDS_SERVICE="rds"
OPEN_SSH_TUNNEL="cloud-tool ssh-tunnel -r 5432 -I -c "
CLOSE_SSH_TUNNEL="cloud-tool ssh-tunnel --close -r 5432 -I -c "

##
# BEGIN CLASS OPData
##
class OPData:

    resource = None
    schema = None

    def __init__(self, schema, resource):
        self.schema = schema
        self.resource = resource

# END CLASS OPData

##
# BEGIN CLASS AWSData
##
class AWSData:

    db = None
    schema = None

    def __init__(self, schema, db):
        self.schema = schema
        self.db = db

# END CLASS AWSData

##
# BEGIN DEF handle_psycopg_exception
# define a function that handles and parses psycopg exceptions
##
def handle_psycopg_exception(err):
    # get details about the exception
    err_type, err_obj, traceback = sys.exc_info()

    # get the line number when exception occured
    line_num = traceback.tb_lineno

    # print the connect() error
    print ("\npsycopg ERROR:", err, "on line number:", line_num)
    print ("psycopg traceback:", traceback, "-- type:", err_type)

    # psycopg extensions.Diagnostics object attribute
    print ("\nextensions.Diagnostics:", err.diag)

    # print the pgcode and pgerror exceptions
    print ("pgerror:", err.pgerror)
    print ("pgcode:", err.pgcode, "\n")

# END print_psycopg_exception()

##
# BEGIN DEF handle_aws_error(e)
##
def handle_aws_error(e):
    if e.response['Error']['Code'] == 'ResourceNotFoundException':
        print("The requested secret resource was not found")
    elif e.response['Error']['Code'] == 'InvalidRequestException':
        print("The request was invalid due to:", e)
    elif e.response['Error']['Code'] == 'InvalidParameterException':
        print("The request had invalid params:", e)
    elif e.response['Error']['Code'] == 'DecryptionFailure':
        print("The requested secret can't be decrypted using the provided KMS key:", e)
    elif e.response['Error']['Code'] == 'InternalServiceError':
        print("An error occurred on service side:", e)
    else: 
        print("An error occurred on service side:", e)

# END DEF handle_aws_error(e)


##
# BEGIN DEF
# return client for specified AWS region and service
##
def get_client(regionName, serviceName):
    session = boto3.session.Session()

    client = session.client(service_name=SM_SERVICE, region_name=USE_REGION)
    
    return client

# END DEF get_client


##
# BEGIN DEF
# get_resource_dict()
##
def get_resource_dict():

    #filename = "C:\\AWS\\py\\qc_wmd.xlsx"
    filename = "qa_gs.xlsx"

    wb = openpyxl.load_workbook(filename)

    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row
    aws_res_tag = 'tr:db-resource-name'
    leg_res_tag = 'tr:legacy-db-resource-name'

    tag_list = []
    resource_dict = {}

    for i in range(13, max_row):
        primary_tags = ws.cell(row = i, column = 20).value
        secondary_tags = ws.cell(row = i, column = 21).value
      
      #print(type(primary_tags), primary_tags)
      #print(type(secondary_tags), secondary_tags)

        if(primary_tags):
            tag_list = primary_tags.split('\n')
            key = ''
            value = ''

            for tag in tag_list:
                tag = tag.strip()
                tag = tag.split('=')

                if(tag[0] == leg_res_tag):
                    key = tag[1]
                elif (tag[0] == aws_res_tag):
                    value = tag[1]
            
            resource_dict[key] = value
        
    if(secondary_tags):
        tag_list = secondary_tags.split('\n')
        key = ''
        value = ''

        for tag in tag_list:
            tag = tag.strip()
            tag = tag.split('=')

            if(tag[0] == leg_res_tag):
               key = tag[1]
            elif (tag[0] == aws_res_tag):
               value = tag[1]

        resource_dict[key] = value
        
    return resource_dict

# END DEF get_resource_dict()

